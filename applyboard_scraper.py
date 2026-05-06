from __future__ import annotations

import csv
import datetime as dt
import os
import re
import sys
import time
from dataclasses import dataclass
from getpass import getpass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
import argparse

import pandas as pd
from dotenv import load_dotenv
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    InvalidSessionIdException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

_DEST_COUNTRY_CACHE: Dict[str, str] = {}
_MYSQL_SINK: Optional["MySQLSink"] = None


def _parse_php_db_config(db_php_path: Path) -> Optional[Dict[str, str]]:
    """
    Parse a simple PHP config like:
      $host = 'localhost';
      $db   = 'visaeofi_cyprus';
      $user = 'root';
      $pass = '';
    """
    try:
        txt = db_php_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None

    def pick(var: str) -> str:
        m = re.search(rf"\${re.escape(var)}\s*=\s*['\"]([^'\"]*)['\"]\s*;", txt)
        return (m.group(1) if m else "").strip()

    host = pick("host")
    db = pick("db") or pick("dbname")  # db.php uses $dbname
    user = pick("user")
    pw = pick("pass")
    port_s = pick("port")
    port = int(port_s) if port_s.isdigit() else 3306
    if not host or not db or not user:
        return None
    return {"host": host, "db": db, "user": user, "password": pw, "port": str(port)}


class MySQLSink:
    def __init__(self, host: str, db: str, user: str, password: str, port: int = 3306) -> None:
        import mysql.connector  # type: ignore

        self._mysql = mysql.connector
        self.host = host
        self.db = db
        self.user = user
        self.password = password
        self.port = int(port or 3306)
        self._conn = None

    def connect(self) -> None:
        if self._conn is not None:
            return
        self._conn = self._mysql.connect(
            host=self.host,
            port=self.port,
            user=self.user,
            password=self.password,
            database=self.db,
            autocommit=True,
        )
        cur = self._conn.cursor()
        cur.execute("SET NAMES utf8mb4")
        cur.close()

    def close(self) -> None:
        try:
            if self._conn is not None:
                self._conn.close()
        except Exception:
            pass
        self._conn = None

    def ensure_table(self, table: str = "applyboard_students") -> None:
        self.connect()
        assert self._conn is not None
        cur = self._conn.cursor()
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS `{table}` (
              `Student ID` VARCHAR(32) NOT NULL,
              `Applicant Name` VARCHAR(255) NULL,
              `Registration Date` VARCHAR(64) NULL,
              `Applicant Email` VARCHAR(255) NULL,
              `Education Level` VARCHAR(255) NULL,
              `Destination/Country` VARCHAR(255) NULL,
              `Target University` TEXT NULL,
              `Target Program` TEXT NULL,
              `Target Intake` TEXT NULL,
              `WhatsApp Num` VARCHAR(64) NULL,
              `Country` VARCHAR(128) NULL,
              `City` VARCHAR(128) NULL,
              `Status App` VARCHAR(512) NULL,
              `In-take` VARCHAR(128) NULL,
              `Graduation / Notes` TEXT NULL,
              `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`Student ID`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
        )
        self._migrate_legacy_applyboard_columns(cur, table)
        cur.close()

    def _migrate_legacy_applyboard_columns(self, cur: Any, table: str) -> None:
        """
        Older versions used snake_case columns. Rename in-place so DB headers match EXCEL_COLUMNS.
        Safe to run repeatedly.
        """
        pairs = [
            ("student_id", "Student ID", "`Student ID` VARCHAR(32) NOT NULL"),
            ("applicant_name", "Applicant Name", "`Applicant Name` VARCHAR(255) NULL"),
            ("registration_date", "Registration Date", "`Registration Date` VARCHAR(64) NULL"),
            ("applicant_email", "Applicant Email", "`Applicant Email` VARCHAR(255) NULL"),
            ("education_level", "Education Level", "`Education Level` VARCHAR(255) NULL"),
            ("destination_country", "Destination/Country", "`Destination/Country` VARCHAR(255) NULL"),
            ("target_university", "Target University", "`Target University` TEXT NULL"),
            ("target_program", "Target Program", "`Target Program` TEXT NULL"),
            ("target_intake", "Target Intake", "`Target Intake` TEXT NULL"),
            ("whatsapp_num", "WhatsApp Num", "`WhatsApp Num` VARCHAR(64) NULL"),
            ("country", "Country", "`Country` VARCHAR(128) NULL"),
            ("city", "City", "`City` VARCHAR(128) NULL"),
            ("status_app", "Status App", "`Status App` VARCHAR(512) NULL"),
            ("intake_label", "In-take", "`In-take` VARCHAR(128) NULL"),
        ]

        def has_col(name: str) -> bool:
            cur.execute(
                """
                SELECT COUNT(*) FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = %s
                """,
                (table, name),
            )
            row = cur.fetchone()
            return bool(row and row[0])

        for old, new, ddl in pairs:
            if has_col(new):
                continue
            if has_col(old):
                cur.execute(f"ALTER TABLE `{table}` CHANGE COLUMN `{old}` {ddl}")

        # New template column (older schemas stored profile_url separately; Excel uses Graduation / Notes)
        if not has_col("Graduation / Notes"):
            cur.execute(f"ALTER TABLE `{table}` ADD COLUMN `Graduation / Notes` TEXT NULL")
        if has_col("profile_url"):
            try:
                cur.execute(f"ALTER TABLE `{table}` DROP COLUMN `profile_url`")
            except Exception:
                pass

    def upsert_student(self, row: "StudentRow", table: str = "applyboard_students") -> None:
        """
        Upsert rule:
        - Only update a field if the new value is non-empty AND different from existing.
        - Especially for status_app: never overwrite with empty.
        """
        self.connect()
        assert self._conn is not None

        data = {
            "Student ID": (row.student_id or "").strip(),
            "Applicant Name": f"{row.first_name} {row.last_name}".strip(),
            "Registration Date": (row.registration_date or "").strip(),
            "Applicant Email": (row.student_email or "").strip(),
            "Education Level": (row.education or "").strip(),
            "Destination/Country": (row.destination_country or "").strip(),
            "Target University": (row.target_university or "").strip(),
            "Target Program": (row.target_program or "").strip(),
            "Target Intake": (row.target_intake or "").strip(),
            "WhatsApp Num": (row.phone or "").strip(),
            "Country": (row.nationality or "").strip(),
            "City": (row.city or "").strip(),
            "Status App": (row.status_app or "").strip(),
            "In-take": (row.intake_label or "").strip(),
            "Graduation / Notes": (row.graduation_notes or "").strip(),
        }

        cols = list(data.keys())
        placeholders = ", ".join(["%s"] * len(cols))
        col_list = ", ".join([f"`{c}`" for c in cols])

        # Conditional update: keep old if incoming is '' or same
        updates = []
        for c in cols:
            if c == "Student ID":
                continue
            updates.append(f"`{c}` = IF(VALUES(`{c}`) IS NULL OR VALUES(`{c}`) = '' OR VALUES(`{c}`) = `{c}`, `{c}`, VALUES(`{c}`))")

        sql = (
            f"INSERT INTO `{table}` ({col_list}) VALUES ({placeholders}) "
            + "ON DUPLICATE KEY UPDATE "
            + ", ".join(updates)
        )

        cur = self._conn.cursor()
        cur.execute(sql, [data[c] for c in cols])
        cur.close()


def infer_destination_country_from_university(university_name: str) -> str:
    """
    Uses OPENAI_API_KEY (if available) to map university/school -> destination country.
    Cached in-memory for the run. Never logs the API key.
    """
    u = (university_name or "").strip()
    if not u:
        return ""
    if u in _DEST_COUNTRY_CACHE:
        return _DEST_COUNTRY_CACHE[u]

    key = (os.getenv("OPENAI_API_KEY") or "").strip()
    if not key:
        _DEST_COUNTRY_CACHE[u] = ""
        return ""

    try:
        from openai import OpenAI  # type: ignore

        client = OpenAI(api_key=key)
        prompt = (
            "Return ONLY the destination country for this university/school name.\n"
            "If unsure, return an empty string.\n\n"
            f"University/School: {u}\n"
        )
        resp = client.responses.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            input=prompt,
        )
        out = (getattr(resp, "output_text", "") or "").strip()
        # Keep a simple plain country name (no punctuation/newlines)
        out = re.sub(r"[\r\n]+", " ", out).strip()
        out = re.sub(r"^[\"'`]+|[\"'`]+$", "", out).strip()
        if len(out) > 64:
            out = out[:64].strip()
        _DEST_COUNTRY_CACHE[u] = out
        return out
    except Exception:
        _DEST_COUNTRY_CACHE[u] = ""
        return ""


# ApplyBoard Identity (recommended — Email/Phone tiles, same as id.applyboard.com). Override with APPLYBOARD_LOGIN_URL.
# Full OAuth URLs with state=/nonce= expire; use the bare origin unless you know you need query params.
DEFAULT_LOGIN_ENTRY_URL = "https://id.applyboard.com/"
LOGIN_URL = DEFAULT_LOGIN_ENTRY_URL  # legacy alias
OKTA_LOGIN_URL = "https://accounts.applyboard.com/oauth2/default/v1/authorize"


def effective_login_url(cli_override: str = "") -> str:
    """CLI --login-url wins, then APPLYBOARD_LOGIN_URL, then default. Call after load_dotenv()."""
    u = (cli_override or "").strip()
    if u:
        return u
    u = (os.getenv("APPLYBOARD_LOGIN_URL") or "").strip()
    return u if u else DEFAULT_LOGIN_ENTRY_URL


# Primary Students destination requested by the user.
STUDENTS_URL = "https://www.applyboard.com/students"


def effective_mysql_db_php_path(*, base_dir: Path, cli_value: str) -> Path:
    """
    Decide which db.php to use for MySQL credentials.
    Priority:
    - CLI --mysql-db-php (if provided)
    - env APPLYBOARD_MYSQL_DB_PHP
    - ./db.php next to this script (useful on VPS: /var/www/.../db.php)
    - Windows default XAMPP db.php (keeps your local workflow)
    """
    v = (cli_value or "").strip()
    if v:
        return Path(v)
    v = (os.getenv("APPLYBOARD_MYSQL_DB_PHP") or "").strip()
    if v:
        return Path(v)
    local = base_dir / "db.php"
    if local.exists():
        return local
    if sys.platform == "win32":
        return Path(r"C:\xampp\htdocs\parrot_mis\db.php")
    return local

# Set after first successful navigation to a working Students list route for this session.
_EFFECTIVE_STUDENTS_URL: Optional[str] = None

EMAIL_XPATH = "//input[@type='email']"
NEXT_BTN_XPATH = "//button[contains(., 'Next')]"
PASSWORD_XPATH = "//input[@type='password']"
VERIFY_BTN_XPATH = "//button[contains(., 'Verify')]"

# Okta login commonly uses these IDs/names (varies by tenant/theme).
# NOTE: Only used after we know we're on an Okta-hosted page — includes broad selectors.
OKTA_USERNAME_LOCATORS = [
    (By.ID, "okta-signin-username"),
    (By.NAME, "username"),
    (By.NAME, "identifier"),
    (By.CSS_SELECTOR, "input[type='email']"),
    (By.CSS_SELECTOR, "input[type='text']"),
]

# Strict markers only — partners ApplyBoard email steps also use type=email/text; those must NOT trigger Okta.
OKTA_FLOW_MARKERS = [
    (By.ID, "okta-signin-username"),
    (By.CSS_SELECTOR, "form[data-se='o-form']"),
    (By.CSS_SELECTOR, "div.okta-sign-in-header"),
    (By.CSS_SELECTOR, "div[id='okta-sign-in']"),
]
OKTA_PASSWORD_LOCATORS = [
    (By.ID, "okta-signin-password"),
    (By.NAME, "password"),
    (By.CSS_SELECTOR, "input[type='password']"),
]
OKTA_SUBMIT_LOCATORS = [
    (By.ID, "okta-signin-submit"),
    (By.CSS_SELECTOR, "input[type='submit']"),
    (By.XPATH, "//button[@type='submit']"),
    (By.XPATH, "//button[contains(., 'Sign in')]"),
    (By.XPATH, "//button[contains(., 'Verify')]"),
    (By.XPATH, "//button[contains(., 'Next')]"),
]

# Partners splash / pre-login page (as seen in your screenshot)
GO_TO_APPLYBOARD_LOCATORS = [
    (By.XPATH, "//button[contains(., 'Go to ApplyBoard')]"),
    (By.XPATH, "//a[contains(., 'Go to ApplyBoard')]"),
    (By.XPATH, "//*[self::button or self::a][contains(., 'Go to ApplyBoard')]"),
]

# Student-side auth pages sometimes show a registration form first.
LOGIN_LINK_LOCATORS = [
    (By.XPATH, "//a[contains(., 'Log in') or contains(., 'Log In') or contains(., 'Login')]"),
    (By.XPATH, "//button[contains(., 'Log in') or contains(., 'Log In') or contains(., 'Login')]"),
]

# Prefer specific email fields first — generic input[type=text] can exist on the Email/Phone *picker* and
# would skip clicking "Email Address" if matched too early.
FLEX_EMAIL_LOCATORS_PRIMARY = [
    (By.XPATH, "//input[@type='email']"),
    (By.XPATH, "//input[contains(@placeholder, 'Enter your email') or contains(@placeholder, 'enter your email')]"),
    (By.XPATH, "//input[contains(@placeholder, 'email') or contains(@placeholder, 'Email')]"),
    (By.XPATH, "//input[contains(@aria-label, 'email') or contains(@aria-label, 'Email')]"),
    (By.CSS_SELECTOR, "input[name='email']"),
    (By.CSS_SELECTOR, "input[id*='email' i]"),
    (By.CSS_SELECTOR, "input[autocomplete='email']"),
    (By.CSS_SELECTOR, "input[inputmode='email']"),
    (By.XPATH, "//label[contains(., 'Email')]/following::input[1]"),
]
FLEX_EMAIL_LOCATORS_FALLBACK = [
    (By.CSS_SELECTOR, "input[type='text']"),
]

FLEX_EMAIL_LOCATORS = FLEX_EMAIL_LOCATORS_PRIMARY + FLEX_EMAIL_LOCATORS_FALLBACK

# Partners login: "Continue with email or phone" → click Email Address before the email field exists.
# Tiles are often <div role="button"> or clickable divs, not <button>; text may sit in child spans.
EMAIL_SIGN_IN_METHOD_LOCATORS = [
    (By.XPATH, "//button[contains(normalize-space(.), 'Email Address')]"),
    (By.XPATH, "//button[contains(., 'Email Address')]"),
    (By.XPATH, "//button[.//span[contains(., 'Email Address')]]"),
    (By.XPATH, "//span[contains(normalize-space(.), 'Email Address')]/ancestor::*[self::button or @role='button'][1]"),
    (By.XPATH, "//span[contains(., 'Email Address')]/ancestor::div[@role='button'][1]"),
    (By.XPATH, "//div[@role='button'][contains(., 'Email Address')]"),
    (By.XPATH, "//div[@role='button'][contains(., 'Email')][contains(., 'Address')]"),
    (By.XPATH, "//*[@role='button'][contains(normalize-space(.), 'Email Address')]"),
    (By.XPATH, "//a[contains(., 'Email Address')]"),
    # Div "cards" (cursor/tabindex) — exclude Phone row.
    (By.XPATH, "//div[@tabindex][contains(., 'Email')][contains(., 'Address')][not(contains(., 'Phone Number'))]"),
    (By.XPATH, "//div[contains(@class,'cursor-pointer')][contains(., 'Email')][not(contains(., 'Phone'))]"),
    (By.XPATH, "//*[(@role='button') or (self::button)][contains(., 'Email')][contains(., 'Address')][not(contains(., 'Phone Number'))]"),
]

FLEX_SUBMIT_LOCATORS = [
    (By.XPATH, "//button[contains(., 'Next')]"),
    (By.XPATH, "//button[contains(., 'Verify')]"),
    (By.XPATH, "//button[contains(., 'Continue')]"),
    (By.XPATH, "//button[contains(., 'Sign in') or contains(., 'Sign In')]"),
    (By.XPATH, "//button[contains(., 'Log in') or contains(., 'Log In') or contains(., 'Login')]"),
    (By.CSS_SELECTOR, "button[type='submit']"),
]

TABLE_ROW_SELECTORS = [
    "table tbody tr",
    "table tr",  # fallback if tbody isn't used
    "div[role='row']",  # some React grids
]

OUTPUT_XLSX = "applyboard_students.xlsx"
OUTPUT_CSV = "applyboard_students.csv"

# Excel output order (matches your template). Empty fields stay blank until we scrape more pages.
EXCEL_COLUMNS: List[str] = [
    "S/N",
    "Applicant Name",
    "Student ID",
    "Registration Date",
    "Applicant Email",
    "Education Level",
    "Destination/Country",
    "Target University",
    "Target Program",
    "Target Intake",
    "WhatsApp Num",
    "Country",
    "City",
    "Status App",
    "In-take",
    "Graduation / Notes",
]

DEFAULT_TIMEOUT_S = 30
CLICK_RETRIES = 4


@dataclass
class StudentRow:
    student_id: str = ""
    student_email: str = ""
    first_name: str = ""
    last_name: str = ""
    nationality: str = ""
    recruitment_partner: str = ""
    recruiter_type: str = ""
    education: str = ""
    phone: str = ""
    registration_date: str = ""
    profile_url: str = ""
    page_number: int = 0
    row_number: int = 0
    # Template columns (optional; fill later if you add Applications / address scrape)
    destination_country: str = ""
    target_university: str = ""
    target_program: str = ""
    target_intake: str = ""
    city: str = ""
    status_app: str = ""
    intake_label: str = ""
    graduation_notes: str = ""

    def to_excel_row(self, sn: int) -> Dict[str, Any]:
        applicant = f"{self.first_name} {self.last_name}".strip()
        return {
            "S/N": sn,
            "Applicant Name": applicant,
            "Student ID": self.student_id,
            "Registration Date": self.registration_date,
            "Applicant Email": self.student_email,
            "Education Level": self.education,
            "Destination/Country": self.destination_country,
            "Target University": self.target_university,
            "Target Program": self.target_program,
            "Target Intake": self.target_intake,
            "WhatsApp Num": self.phone,
            "Country": self.nationality,
            "City": self.city,
            "Status App": self.status_app,
            "In-take": self.intake_label,
            "Graduation / Notes": self.graduation_notes,
        }


def now_stamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_text(el: Optional[WebElement]) -> str:
    if el is None:
        return ""
    try:
        return (el.text or "").strip()
    except Exception:
        return ""


def ensure_dirs(base_dir: Path) -> Tuple[Path, Path]:
    out_dir = base_dir
    err_dir = base_dir / "errors"
    err_dir.mkdir(parents=True, exist_ok=True)
    return out_dir, err_dir


def screenshot_on_error(driver: webdriver.Chrome, err_dir: Path, name: str) -> Optional[Path]:
    try:
        path = err_dir / f"{now_stamp()}_{name}.png"
        driver.save_screenshot(str(path))
        return path
    except (InvalidSessionIdException, WebDriverException):
        return None
    except Exception:
        return None


def _safe_quit_driver(driver: Optional[webdriver.Chrome]) -> None:
    """Avoid crashing on quit when Chrome was already closed or the session died."""
    if driver is None:
        return
    try:
        driver.quit()
    except (InvalidSessionIdException, WebDriverException):
        pass
    except Exception:
        pass


def wait_for_page(driver: webdriver.Chrome, timeout_s: int = DEFAULT_TIMEOUT_S) -> None:
    WebDriverWait(driver, timeout_s).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def _safe_driver_get(driver: webdriver.Chrome, url: str, *, page_load_timeout_s: int = 180) -> None:
    """
    Heavy SPA navigations can exceed default timeouts and occasionally stress Chrome.
    Stop loading on timeout but keep the session so Selenium can continue.
    """
    try:
        driver.set_page_load_timeout(page_load_timeout_s)
    except Exception:
        pass
    try:
        driver.get(url)
    except TimeoutException:
        try:
            driver.execute_script("window.stop();")
        except Exception:
            pass
    except InvalidSessionIdException:
        raise


def click_with_retry(
    driver: webdriver.Chrome,
    element: WebElement,
    timeout_s: int = DEFAULT_TIMEOUT_S,
    retries: int = CLICK_RETRIES,
) -> None:
    last_exc: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            WebDriverWait(driver, timeout_s).until(EC.visibility_of(element))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.15)
            element.click()
            return
        except (StaleElementReferenceException, ElementClickInterceptedException, WebDriverException) as exc:
            last_exc = exc
            time.sleep(min(1.2 * attempt, 4.0))
    if last_exc:
        raise last_exc


def _cookies_required_interstitial(driver: webdriver.Chrome) -> bool:
    try:
        body = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        return False
    return ("cookies are required" in body) or ("cookies are disabled" in body)


def _page_timed_out(driver: webdriver.Chrome) -> bool:
    try:
        body = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        return False
    return ("the page has timed out" in body) or ("timed out" in body)


def _applyboard_not_found(driver: webdriver.Chrome) -> bool:
    """
    Detect ApplyBoard's styled 404 pages so we can skip bad URLs quickly.
    """
    try:
        body = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body = ""
    try:
        title = (driver.title or "").lower()
    except Exception:
        title = ""

    return ("404" in body and "not found" in body) or ("page not found" in body) or ("404" in title)


def _find_first_present(
    driver: webdriver.Chrome, locators: List[Tuple[str, str]]
) -> Optional[WebElement]:
    for by, value in locators:
        try:
            el = driver.find_element(by, value)
            if el and el.is_displayed() and el.is_enabled():
                return el
        except Exception:
            continue
    return None


def _find_first_visible(driver: webdriver.Chrome, locators: List[Tuple[str, str]]) -> Optional[WebElement]:
    """Like _find_first_present but only requires visibility (method-picker buttons can be odd)."""
    for by, value in locators:
        try:
            el = driver.find_element(by, value)
            if el and el.is_displayed():
                return el
        except Exception:
            continue
    return None


def _find_partner_email_input_in_context(driver: webdriver.Chrome) -> Optional[WebElement]:
    """
    Partners email step: input may be visible before Selenium reports enabled=True.
    Prefer interactable fields; fall back to displayed-only <input>.
    Generic type=text is last — avoid matching hidden/auxiliary fields on the Email vs Phone picker screen.
    """
    for group in (FLEX_EMAIL_LOCATORS_PRIMARY, FLEX_EMAIL_LOCATORS_FALLBACK):
        el = _find_first_present(driver, group)
        if el is not None:
            return el
        for by, value in group:
            try:
                cand = driver.find_element(by, value)
                tag = (cand.tag_name or "").lower()
                if tag != "input":
                    continue
                if cand.is_displayed():
                    return cand
            except Exception:
                continue
    return None


def _find_partner_email_input_any_frame(driver: webdriver.Chrome) -> Optional[WebElement]:
    """Email field may be inside an iframe (same widget as the Email/Phone picker)."""
    driver.switch_to.default_content()
    el = _find_partner_email_input_in_context(driver)
    if el is not None:
        return el
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    for frame in frames:
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            el = _find_partner_email_input_in_context(driver)
            if el is not None:
                return el
        except Exception:
            continue
    driver.switch_to.default_content()
    return None


def _find_visible_email_method_any_frame(driver: webdriver.Chrome) -> Optional[WebElement]:
    driver.switch_to.default_content()
    el = _find_first_visible(driver, EMAIL_SIGN_IN_METHOD_LOCATORS)
    if el is not None:
        return el
    for frame in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            el = _find_first_visible(driver, EMAIL_SIGN_IN_METHOD_LOCATORS)
            if el is not None:
                return el
        except Exception:
            continue
    driver.switch_to.default_content()
    return None


def _fill_input(driver: webdriver.Chrome, el: WebElement, value: str) -> None:
    """
    Robustly fills an <input> that may be present but not yet interactable.
    """
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
    try:
        el.click()
    except Exception:
        try:
            driver.execute_script("arguments[0].focus();", el)
        except Exception:
            pass
    time.sleep(0.15)

    try:
        el.clear()
    except Exception:
        # Some React/Okta themed inputs block clear(); fall back to Ctrl+A Backspace.
        try:
            el.send_keys("\u0001")  # Ctrl+A
            el.send_keys("\b")
        except Exception:
            pass
    el.send_keys(value)


def _find_first_present_in_any_frame(
    driver: webdriver.Chrome, locators: List[Tuple[str, str]]
) -> Optional[WebElement]:
    driver.switch_to.default_content()
    el = _find_first_present(driver, locators)
    if el is not None:
        return el

    frames = driver.find_elements(By.TAG_NAME, "iframe")
    for frame in frames:
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            el = _find_first_present(driver, locators)
            if el is not None:
                return el
        except Exception:
            continue
    driver.switch_to.default_content()
    return None


def _find_submit_continue_any_frame(driver: webdriver.Chrome) -> Optional[WebElement]:
    """Identity flow uses Continue → spinner → password; ensure we find Continue in main DOM or iframes."""
    locators: List[Tuple[str, str]] = [
        (By.XPATH, "//button[contains(normalize-space(.), 'Continue')]"),
        (By.XPATH, "//button[contains(., 'Continue')]"),
        (By.XPATH, "//*[@role='button'][contains(., 'Continue')]"),
    ]
    locators.extend(FLEX_SUBMIT_LOCATORS)
    locators.append((By.XPATH, NEXT_BTN_XPATH))

    driver.switch_to.default_content()
    for by, value in locators:
        try:
            for el in driver.find_elements(by, value):
                try:
                    if el.is_displayed() and el.is_enabled():
                        return el
                except Exception:
                    continue
        except Exception:
            continue

    for frame in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            for by, value in locators:
                try:
                    for el in driver.find_elements(by, value):
                        try:
                            if el.is_displayed() and el.is_enabled():
                                return el
                        except Exception:
                            continue
                except Exception:
                    continue
        except Exception:
            continue
    driver.switch_to.default_content()
    return None


def _wait_for_password_input_any_frame(
    driver: webdriver.Chrome,
    timeout_s: float = 95.0,
) -> Optional[WebElement]:
    """
    After Continue, ApplyBoard shows a spinner on the button; password appears asynchronously.
    Poll until a visible password field exists (any frame).
    """
    deadline = time.time() + timeout_s
    pw_locators = [
        (By.XPATH, PASSWORD_XPATH),
        (By.CSS_SELECTOR, "input[type='password']"),
        (By.CSS_SELECTOR, "input[name='password']"),
        (By.CSS_SELECTOR, "input[autocomplete='current-password']"),
        (By.CSS_SELECTOR, "input[autocomplete='password']"),
    ]
    while time.time() < deadline:
        driver.switch_to.default_content()
        for by, value in pw_locators:
            try:
                for el in driver.find_elements(by, value):
                    try:
                        if el.is_displayed():
                            return el
                    except Exception:
                        continue
            except Exception:
                continue

        for frame in driver.find_elements(By.TAG_NAME, "iframe"):
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                for by, value in pw_locators:
                    try:
                        for el in driver.find_elements(by, value):
                            try:
                                if el.is_displayed():
                                    return el
                            except Exception:
                                continue
                    except Exception:
                        continue
            except Exception:
                continue
        driver.switch_to.default_content()
        time.sleep(0.45)

    driver.switch_to.default_content()
    return None


def _maybe_click_partners_splash(driver: webdriver.Chrome, err_dir: Path) -> None:
    """
    Some accounts land on a Partners splash page (no login inputs) with a
    'Go to ApplyBoard' button. Click it to proceed to the actual login flow.
    """
    try:
        body_txt = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body_txt = ""

    # Heuristic: look for the CTA and/or the 'partners website' heading.
    cta = _find_first_present(driver, GO_TO_APPLYBOARD_LOCATORS)
    if cta is None and "partners website" not in body_txt:
        return

    if cta is None:
        # Try a broader search in case text is present but selector differs.
        try:
            cta = driver.find_element(By.XPATH, "//*[self::button or self::a][contains(., 'ApplyBoard')]")
        except Exception:
            cta = None

    if cta is None:
        screenshot_on_error(driver, err_dir, "partners_splash_no_cta")
        return

    print("Detected Partners splash page. Clicking 'Go to ApplyBoard'...")
    old_url = driver.current_url or ""
    old_handles = list(driver.window_handles)

    def did_navigate(d: webdriver.Chrome) -> bool:
        try:
            new_url = d.current_url or ""
            if new_url and new_url != old_url:
                return True
            # Some flows open a new tab/window.
            return len(d.window_handles) > len(old_handles)
        except Exception:
            return False

    click_ok = False
    for _ in range(3):
        try:
            click_with_retry(driver, cta, timeout_s=30, retries=2)
            click_ok = True
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", cta)
                click_ok = True
            except Exception:
                click_ok = False

        if click_ok:
            try:
                WebDriverWait(driver, 15).until(did_navigate)
                break
            except Exception:
                # Still stuck; try again.
                time.sleep(0.8)

    # If a new window/tab was opened, switch to the newest handle.
    try:
        new_handles = list(driver.window_handles)
        if len(new_handles) > len(old_handles):
            newest = [h for h in new_handles if h not in old_handles][-1]
            driver.switch_to.window(newest)
    except Exception:
        pass

    # If we still appear to be on the same splash page, record it and return.
    try:
        wait_for_page(driver, timeout_s=20)
    except Exception:
        pass

    try:
        body_txt2 = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body_txt2 = ""

    if "partners website" in body_txt2 and "go to applyboard" in body_txt2:
        screenshot_on_error(driver, err_dir, "partners_splash_still_visible")
        return

    time.sleep(1.0)


def _maybe_switch_register_to_login(
    driver: webdriver.Chrome, err_dir: Path, *, reload_login_url: str = DEFAULT_LOGIN_ENTRY_URL
) -> None:
    """
    If we land on student *registration* instead of partners login, switch to Log in or reload the login URL.
    When ApplyBoard never shows that UI (common), every check here is false and this returns immediately.
    """
    try:
        cur_u = (driver.current_url or "").lower()
        if "applyboard.com" in cur_u and "/students" in cur_u:
            return
    except Exception:
        pass

    try:
        body_txt = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body_txt = ""

    # Login / Identity modal — same page often mentions "Create account" or footer links; must not reload.
    login_modal_markers = (
        "welcome back",
        "sign in to continue your journey",
        "continue with email or phone",
        "please choose the same method you used to sign up",
    )
    if any(m in body_txt for m in login_modal_markers):
        return

    # Do not key off "already have an account" — it appears on login-adjacent copy and caused false reloads.
    legacy = "register as a student" in body_txt
    # Newer registration wizard (tabs Email/Phone, "Continue registration") — not used by every tenant.
    wizard = ("continue registration" in body_txt) or ("choose your preferred registration method" in body_txt)

    if not legacy and not wizard:
        return

    escape_locators: List[Tuple[str, str]] = list(LOGIN_LINK_LOCATORS) + [
        (By.XPATH, "//a[contains(., 'Already have an account')]"),
        (By.XPATH, "//a[contains(., 'Already have')]"),
    ]

    link = _find_first_present(driver, escape_locators)
    if link is None:
        link = _find_first_visible(driver, escape_locators)

    if link is not None:
        print("Detected registration/onboarding page — switching to Log in...", flush=True)
        try:
            click_with_retry(driver, link, timeout_s=20, retries=2)
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", link)
            except Exception:
                screenshot_on_error(driver, err_dir, "register_page_login_click_failed")
                link = None

        if link is not None:
            time.sleep(1.2)
            try:
                wait_for_page(driver, timeout_s=25)
            except Exception:
                pass
            return

    print("Registration UI detected but no Log in control — reloading login entry URL...", flush=True)
    screenshot_on_error(driver, err_dir, "register_fallback_reload_login")
    _safe_driver_get(driver, reload_login_url)
    try:
        wait_for_page(driver, timeout_s=30)
    except Exception:
        pass


def extract_phone(driver: webdriver.Chrome) -> str:
    # 1) Look for tel: links / visible +<countrycode><number>
    try:
        for a in driver.find_elements(By.CSS_SELECTOR, "a[href]"):
            href = (a.get_attribute("href") or "").strip().lower()
            txt = (a.text or "").strip()
            if href.startswith("tel:"):
                val = href.replace("tel:", "").strip()
                if val:
                    return val
            if txt.startswith("+") and re.search(r"\d{6,}", txt):
                return re.sub(r"\s+", " ", txt).strip()
    except Exception:
        pass

    # 2) Regex search in visible text and page source (some UIs hide text nodes)
    try:
        body_txt = driver.find_element(By.TAG_NAME, "body").text or ""
        m = re.search(r"(\+\d[\d\s\-]{6,}\d)", body_txt.replace("\n", " "))
        if m:
            return re.sub(r"\s+", " ", m.group(1)).strip()
    except Exception:
        pass
    try:
        src = driver.page_source or ""
        m = re.search(r"(\+\d[\d\s\-]{6,}\d)", src)
        if m:
            return re.sub(r"\s+", " ", m.group(1)).strip()
    except Exception:
        pass

    # 2) Look for an input under "Phone Number" label (common on Profile tab)
    try:
        phone_input = driver.find_element(
            By.XPATH,
            "//label[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'phone')]/following::input[1]",
        )
        number = (phone_input.get_attribute("value") or "").strip()
        if number:
            # country code may be in a nearby element (e.g. +86)
            try:
                wrapper = phone_input.find_element(By.XPATH, "ancestor::*[self::div or self::section][1]")
                wtxt = (wrapper.text or "").replace("\n", " ")
                m = re.search(r"(\+\d{1,4})", wtxt)
                if m:
                    return f"{m.group(1)} {number}".strip()
            except Exception:
                pass
            return number
    except Exception:
        pass

    # 3) Fallback: any input[type=tel]
    try:
        tel_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='tel']")
        for ti in tel_inputs:
            v = (ti.get_attribute("value") or "").strip()
            if v and re.search(r"\d{6,}", v):
                return v
    except Exception:
        pass

    # 4) Last resort: any input whose id/name suggests phone
    try:
        candidates = driver.find_elements(
            By.CSS_SELECTOR,
            "input[name*='phone' i], input[id*='phone' i], input[autocomplete*='tel' i]",
        )
        for c in candidates:
            v = (c.get_attribute("value") or "").strip()
            if v and re.search(r"\d{6,}", v):
                return v
    except Exception:
        pass

    return ""


def extract_registration_date(driver: webdriver.Chrome) -> str:
    try:
        body_txt = driver.find_element(By.TAG_NAME, "body").text or ""
    except Exception:
        return ""

    # robust regex: Registration date: Apr 30, 2026
    m = re.search(r"registration date\s*:\s*([A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})", body_txt, re.I)
    if m:
        return m.group(1).strip()

    # fallback: line-based
    for line in [ln.strip() for ln in body_txt.splitlines() if ln.strip()]:
        if "registration date" in line.lower():
            if ":" in line:
                return line.split(":", 1)[1].strip()
            return line
    return ""


def _click_email_method_choice(driver: webdriver.Chrome, el: WebElement) -> None:
    """
    ApplyBoard email/phone tiles are often React divs; native click may fail — use JS click as fallback.
    """
    last_exc: Optional[Exception] = None
    for attempt in range(1, 4):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
            time.sleep(0.2)
            el.click()
            return
        except StaleElementReferenceException:
            raise
        except Exception as exc:
            last_exc = exc
            try:
                driver.execute_script("arguments[0].click();", el)
                return
            except Exception:
                try:
                    driver.execute_script(
                        "arguments[0].dispatchEvent(new MouseEvent('click', {bubbles: true, cancelable: true, view: window}));",
                        el,
                    )
                    return
                except Exception as exc2:
                    last_exc = exc2
            time.sleep(0.35)
    if last_exc:
        raise last_exc


def _maybe_choose_email_login_method(driver: webdriver.Chrome, err_dir: Path) -> bool:
    """
    New ApplyBoard flow: pick Email Address vs Phone before the email input is shown.
    Returns True if a method button was clicked.
    """
    el = _find_visible_email_method_any_frame(driver)
    if el is None:
        driver.switch_to.default_content()
        return False
    try:
        print("Choosing Email Address sign-in method...", flush=True)
        _click_email_method_choice(driver, el)
        time.sleep(1.2)
        driver.switch_to.default_content()
        return True
    except Exception:
        screenshot_on_error(driver, err_dir, "email_method_click_failed")
        driver.switch_to.default_content()
        return False


def login(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    email: str,
    password: str,
    err_dir: Path,
    login_entry_url: Optional[str] = None,
) -> None:
    entry = (login_entry_url or "").strip() or effective_login_url("")
    if "state=" in entry and "nonce=" in entry:
        print(
            "[WARN] Login URL contains OAuth state/nonce parameters — they expire quickly. "
            "Prefer APPLYBOARD_LOGIN_URL=https://id.applyboard.com/ (no query string).",
            flush=True,
        )
    print(f"Opening ApplyBoard login -> {entry}", flush=True)
    driver.get(entry)
    wait_for_page(driver)
    # Identity SPA often renders Loading before Email vs Phone tiles (id.applyboard.com).
    if "id.applyboard.com" in (driver.current_url or "").lower():
        time.sleep(2.5)

    _maybe_click_partners_splash(driver, err_dir)
    _maybe_switch_register_to_login(driver, err_dir, reload_login_url=entry)
    try:
        wait_for_page(driver)
    except Exception:
        pass

    # If partners login redirects into Okta, handle that too.
    current = driver.current_url or ""
    is_okta = ("accounts.applyboard.com" in current) or current.startswith(OKTA_LOGIN_URL)

    if _cookies_required_interstitial(driver):
        screenshot_on_error(driver, err_dir, "cookies_required")
        raise TimeoutException(
            "Login page shows 'Cookies are required'. Enable cookies in Chrome and disable strict privacy extensions."
        )

    # If we're still not sure which login form is showing, detect real Okta UI only (not partners email inputs).
    if not is_okta:
        cur_url = driver.current_url or ""
        if "accounts.applyboard.com" in cur_url:
            is_okta = True
        elif _find_first_present_in_any_frame(driver, OKTA_FLOW_MARKERS) is not None:
            is_okta = True

    if not is_okta:
        # Generic non-Okta login form (partners or student site)
        print("Entering email...")
        driver.switch_to.default_content()
        # This screen is only tiles (Email / Phone) — click Email Address before looking for <input>.
        if _find_visible_email_method_any_frame(driver) is not None:
            print("Email vs Phone picker — clicking Email Address first.", flush=True)
            _maybe_choose_email_login_method(driver, err_dir)
        t0 = time.time()
        email_input: Optional[WebElement] = None
        last_method_try = -100.0
        method_attempts = 0
        while time.time() - t0 < 45 and email_input is None:
            driver.switch_to.default_content()
            email_input = _find_partner_email_input_any_frame(driver)
            if email_input is not None:
                break
            now = time.time()
            # Retry picker occasionally (iframe/slow render); avoid hammering the same button every 350ms.
            if method_attempts < 5 and (now - last_method_try) >= 1.5:
                if _maybe_choose_email_login_method(driver, err_dir):
                    method_attempts += 1
                last_method_try = now
            time.sleep(0.35)

        if email_input is None:
            screenshot_on_error(driver, err_dir, "partners_email_not_found")
            raise TimeoutException("Could not find email input. (Login page UI likely changed.)")

        try:
            _fill_input(driver, email_input, email)
        except Exception:
            screenshot_on_error(driver, err_dir, "email_not_interactable")
            raise

        # If password is on same screen, fill it; otherwise click Continue → wait out spinner → password step.
        pw_locators = [(By.XPATH, PASSWORD_XPATH), (By.CSS_SELECTOR, "input[type='password']")]
        pw_input = _find_first_present_in_any_frame(driver, pw_locators)
        if pw_input is None:
            btn = _find_submit_continue_any_frame(driver)
            if btn is None:
                try:
                    btn = wait.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "//button[contains(., 'Continue') or contains(., 'Next')]")
                        )
                    )
                except Exception:
                    btn = wait.until(EC.element_to_be_clickable((By.XPATH, NEXT_BTN_XPATH)))
            print("Continuing to password step (waiting if button shows a spinner)...", flush=True)
            try:
                click_with_retry(driver, btn)
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
            pw_input = _wait_for_password_input_any_frame(driver, timeout_s=95.0)

        if pw_input is None:
            screenshot_on_error(driver, err_dir, "password_not_found")
            raise TimeoutException("Could not find password input.")

        print("Entering password...")
        try:
            _fill_input(driver, pw_input, password)
        except Exception:
            screenshot_on_error(driver, err_dir, "password_not_interactable")
            raise

        btn = _find_submit_continue_any_frame(driver)
        if btn is None:
            btn = _find_first_present_in_any_frame(driver, FLEX_SUBMIT_LOCATORS)
        if btn is None:
            try:
                btn = wait.until(EC.element_to_be_clickable((By.XPATH, VERIFY_BTN_XPATH)))
            except Exception:
                btn = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//button[contains(., 'Sign in') or contains(., 'Verify')]")
                    )
                )
        try:
            click_with_retry(driver, btn)
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
    else:
        # Okta flow (accounts.applyboard.com)
        print("Detected Okta login. Entering username/email...")
        t0 = time.time()
        username_input: Optional[WebElement] = None
        while time.time() - t0 < 60:
            if _cookies_required_interstitial(driver):
                screenshot_on_error(driver, err_dir, "cookies_required_okta")
                raise TimeoutException(
                    "Okta page shows 'Cookies are required'. Enable cookies in Chrome and disable strict privacy extensions."
                )
            username_input = _find_first_present_in_any_frame(driver, OKTA_USERNAME_LOCATORS)
            if username_input is not None:
                break
            time.sleep(0.4)

        if username_input is None:
            screenshot_on_error(driver, err_dir, "okta_username_not_found")
            raise TimeoutException("Could not find Okta username/email input.")

        try:
            _fill_input(driver, username_input, email)
        except Exception:
            screenshot_on_error(driver, err_dir, "okta_username_not_interactable")
            raise

        # Some Okta pages are single-step (enter both then submit), others might continue.
        submit = _find_first_present_in_any_frame(driver, OKTA_SUBMIT_LOCATORS)
        if submit is not None:
            try:
                click_with_retry(driver, submit)
            except Exception:
                pass

        print("Entering password...")
        t1 = time.time()
        pw_input: Optional[WebElement] = None
        while time.time() - t1 < 60:
            pw_input = _find_first_present_in_any_frame(driver, OKTA_PASSWORD_LOCATORS)
            if pw_input is not None:
                break
            time.sleep(0.4)

        if pw_input is None:
            screenshot_on_error(driver, err_dir, "okta_password_not_found")
            raise TimeoutException("Could not find Okta password input.")

        try:
            _fill_input(driver, pw_input, password)
        except Exception:
            screenshot_on_error(driver, err_dir, "okta_password_not_interactable")
            raise

        submit = _find_first_present_in_any_frame(driver, OKTA_SUBMIT_LOCATORS)
        if submit is None:
            screenshot_on_error(driver, err_dir, "okta_submit_not_found")
            raise TimeoutException("Could not find Okta submit button.")
        click_with_retry(driver, submit)

    print("Finishing login...")
    # Must leave id.applyboard.com + accounts OAuth — old condition passed while still on Identity (instant "success").
    try:
        WebDriverWait(driver, 150).until(lambda d: _post_login_redirect_settled(d))
    except TimeoutException:
        print(
            "Login redirect slow or MFA may be required. If prompted, finish in the browser window.",
            flush=True,
        )
    time.sleep(5.0)


def _post_login_redirect_settled(d: webdriver.Chrome) -> bool:
    u = (d.current_url or "").lower()
    if "authorize" in u:
        return False
    if "accounts.applyboard.com" in u:
        return False
    if "id.applyboard.com" in u:
        return False
    return "applyboard.com" in u


def _looks_like_identity_email_login_screen(driver: webdriver.Chrome) -> bool:
    """Email/Phone sign-in visible — unauthenticated or session not on www yet."""
    try:
        u = (driver.current_url or "").lower()
        body = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        return False
    if "id.applyboard.com" in u and (
        "welcome back" in body
        or "continue with email or phone" in body
        or "enter your email" in body
    ):
        return True
    if "applyboard.com" in u and "/students" not in u and "welcome back" in body and "enter your email" in body:
        return True
    return False


def _navigate_students_with_session_handoff(driver: webdriver.Chrome, err_dir: Path, target_url: str) -> None:
    """www root first (cookie handoff), then Students; retries if we get bounced to sign-in."""
    for attempt in range(1, 4):
        if attempt > 1:
            print(f"[Nav] Students looked unauthenticated — retry {attempt}/3 after longer wait...", flush=True)
            time.sleep(6.0 * attempt)
        _safe_driver_get(driver, "https://www.applyboard.com/")
        try:
            wait_for_page(driver, timeout_s=45)
        except Exception:
            pass
        time.sleep(2.0)
        _safe_driver_get(driver, target_url)
        try:
            wait_for_page(driver, timeout_s=60)
        except Exception:
            pass
        time.sleep(2.0)
        if not _looks_like_identity_email_login_screen(driver):
            return
        screenshot_on_error(driver, err_dir, f"students_bounced_login_attempt_{attempt}")

    raise TimeoutException(
        "ApplyBoard keeps showing sign-in when opening Students — complete MFA if shown, or try logging in once in "
        "the same browser profile before running the scraper."
    )


def _on_agent_dashboard_not_students(driver: webdriver.Chrome) -> bool:
    u = (driver.current_url or "").lower()
    return "applyboard.com" in u and "/agent" in u and "/students" not in u


def _loose_students_grid_has_data(driver: webdriver.Chrome) -> bool:
    """
    If strict row heuristics miss (DOM quirks), treat table as ready when we see
    student-like rows: 6+ digit id and an email in the same tbody row.
    """
    try:
        for tr in driver.find_elements(By.CSS_SELECTOR, "table tbody tr")[:40]:
            txt = (tr.text or "").replace("\n", " ")
            if re.search(r"\b\d{6,}\b", txt) and "@" in txt:
                return True
    except Exception:
        pass
    return False


def _wait_students_list_ready(driver: webdriver.Chrome, err_dir: Path, timeout_s: int = 180) -> None:
    """
    Wait for the Students SPA to finish (grey spinner) and for **real** student rows (not header/filter).
    Never uses browser Back() — that commonly pops to /agent and causes agent/students flip-flop.
    """
    deadline = time.time() + timeout_s
    agent_redirect_count = 0
    start = time.time()
    while time.time() < deadline:
        if _cookies_required_interstitial(driver):
            screenshot_on_error(driver, err_dir, "students_cookies_required")
            raise TimeoutException("Students page shows 'Cookies are required'.")
        if _page_timed_out(driver):
            screenshot_on_error(driver, err_dir, "students_timed_out")
            raise TimeoutException("Students page timed out.")

        cur = driver.current_url or ""
        if _on_agent_dashboard_not_students(driver):
            if agent_redirect_count < 4:
                print("[Nav] On /agent; opening Students URL...", flush=True)
                _safe_driver_get(driver, STUDENTS_URL)
                wait_for_page(driver)
                agent_redirect_count += 1
            time.sleep(0.5)
            continue

        if "partners.applyboard.com" in cur:
            _maybe_click_partners_splash(driver, err_dir=err_dir)
            try:
                wait_for_page(driver, timeout_s=15)
            except Exception:
                pass

        try:
            if len(get_student_data_rows(driver)) > 0:
                return
            # After a few seconds, accept loose match so we don't timeout while the table is visible.
            if time.time() - start > 12 and _loose_students_grid_has_data(driver):
                print("[Nav] Students table ready (loose check: ID + email in row).", flush=True)
                return
        except Exception:
            pass
        time.sleep(0.5)

    screenshot_on_error(driver, err_dir, "students_table_timeout")
    raise TimeoutException(
        "Students table did not appear after waiting (page may still be loading — try increasing timeout)."
    )


def wait_for_students_table(driver: webdriver.Chrome, wait: WebDriverWait, err_dir: Path) -> None:
    # Use stable www routes only — /agent/students often 404s or redirects and triggers flip-flopping.
    candidate_urls = [
        STUDENTS_URL,
        "https://applyboard.com/students",
        "https://partners.applyboard.com/students",
    ]

    # Post-login we may already be on /students — avoid a second driver.get (reduces Chrome crashes + login bounce).
    try:
        cur = (driver.current_url or "").lower()
        if "applyboard.com" in cur and "/students" in cur and not _applyboard_not_found(driver):
            print(f"Already on Students ({driver.current_url}); waiting for grid...", flush=True)
            if _looks_like_identity_email_login_screen(driver):
                _navigate_students_with_session_handoff(driver, err_dir, STUDENTS_URL)
            _maybe_click_partners_splash(driver, err_dir=err_dir)
            _maybe_switch_register_to_login(
                driver, err_dir=err_dir, reload_login_url=effective_login_url()
            )
            _wait_students_list_ready(driver, err_dir=err_dir, timeout_s=180)
            _set_effective_students_url(STUDENTS_URL)
            return
    except Exception:
        pass

    last_exc: Optional[Exception] = None
    for url in candidate_urls:
        try:
            print(f"Navigating to Students page: {url}")
            _navigate_students_with_session_handoff(driver, err_dir, url)
            wait_for_page(driver)
            if _applyboard_not_found(driver):
                screenshot_on_error(driver, err_dir, "students_404")
                raise TimeoutException(f"Students URL returned a 404-like page: {url}")

            _maybe_click_partners_splash(driver, err_dir=err_dir)
            _maybe_switch_register_to_login(
                driver, err_dir=err_dir, reload_login_url=effective_login_url()
            )

            _wait_students_list_ready(driver, err_dir=err_dir, timeout_s=180)
            _set_effective_students_url(STUDENTS_URL)
            return
        except Exception as exc:
            last_exc = exc
            continue

    if last_exc:
        raise last_exc


def parse_row_cells(cells: List[WebElement]) -> Dict[str, str]:
    # Legacy index-based parsing. Prefer header-based mapping via parse_row_by_headers().
    def cell(i: int) -> str:
        return safe_text(cells[i]) if i < len(cells) else ""

    return {
        "student_id": cell(1),
        "student_email": cell(2),
        "first_name": cell(3),
        "last_name": cell(4),
        "nationality": cell(5),
        "recruitment_partner": cell(6),
        "recruiter_type": cell(7),
        "education": cell(8),
    }


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("...", "").replace("…", "")
    return s


# Secondary lines inside Paid Apps "Status" column (workflow — not the headline status)
_APPLYBOARD_SUBSTATUS_LINES = frozenset(
    {
        "ready to submit",
        "submitted to school",
        "pre-submission",
        "pre submission",
        "post-decision",
        "post decision",
        "admission",
    }
)


def _is_applyboard_substatus_line(line: str) -> bool:
    low = (line or "").strip().lower()
    if low in _APPLYBOARD_SUBSTATUS_LINES:
        return True
    return any(
        x in low
        for x in (
            "ready to submit",
            "submitted to school",
            "pre-submission",
            "post-decision",
        )
    )


def _primary_paid_application_status(raw: str) -> str:
    """
    Paid Applications Status cell often stacks lines, e.g.:
      Rejected
      Ready to Submit
      Submitted to School
    Return the real headline status (Rejected / Accepted / Canceled / Processing / …).
    """
    if not raw or not str(raw).strip():
        return ""
    lines = [ln.strip() for ln in re.split(r"[\r\n]+", str(raw)) if ln.strip()]
    if not lines:
        return ""

    # Exact primary labels (ApplyBoard UI)
    exact_primary = {
        "rejected",
        "canceled",
        "cancelled",
        "accepted",
        "processing",
        "submitted",
        "draft",
        "admitted",
        "deferred",
        "withdrawn",
        "waitlisted",
        "waitlist",
        "denied",
        "pending",
        "under review",
        "conditional offer",
        "offer letter",
    }

    for ln in lines:
        if _is_applyboard_substatus_line(ln):
            continue
        low = ln.lower().strip()
        if low in exact_primary:
            return ln.strip()
        # Short headline lines containing status keywords
        if len(ln) <= 56:
            if re.search(
                r"\b(rejected|cancel(?:led|ed)|accepted|admitted|processing|submitted|draft|deferred|"
                r"withdrawn|withdraw|waitlisted|waitlist|denied|pending|under[\s-]+review)\b",
                low,
            ):
                return ln.strip()

    # Fallback: first line that isn't a known sub-status workflow
    for ln in lines:
        if not _is_applyboard_substatus_line(ln):
            return ln.strip()
    return lines[0].strip()


def _apps_table_column_map(headers: List[Tuple[int, str]]) -> Dict[str, int]:
    """
    Map ApplyBoard application grids with ambiguous duplicate headers (e.g., multiple 'Program' columns).
    """
    m: Dict[str, int] = {}

    def cand(kind: str, idx: int, hdr: str, score: int) -> None:
        cur = m.get(kind)
        cur_score = m.get(f"__score::{kind}", -999)
        if cur is None or score > cur_score:
            m[kind] = idx
            m[f"__score::{kind}"] = score

    program_idxs: List[int] = []
    for i, h in headers:
        if not h:
            continue
        hl = h.lower()
        if hl.startswith("id") and "student" not in hl:
            cand("id", i, h, 50)

        # School / university columns (avoid unrelated uses of the word "college")
        if any(k in hl for k in ("school", "university", "institution", "campus")):
            cand("school", i, h, 80 if "school" in hl else 70)

        if "status" in hl and not any(x in hl for x in ("requirements", "requirement")):
            cand("status", i, h, 70)

        if "intake" in hl:
            cand("intake", i, h, 80)

        if "requirement" in hl:
            cand("requirements", i, h, 75)

        if "program" in hl:
            program_idxs.append(i)

    # Pick the best Program column when multiple exist (often one is actually Requirements).
    best_pi = -1
    best_pscore = -999
    for i in program_idxs:
        hh = ""
        for ii, hx in headers:
            if ii == i:
                hh = hx
                break
        hl = hh.lower()
        score = 80 if hl.strip() == "program" else 60
        if any(k in hl for k in ("requirement", "needed", "prerequisite", "readiness", "pre-req", "prereq")):
            score -= 50
        if "target" in hl or "selected" in hl or "primary" in hl:
            score += 10
        if score > best_pscore:
            best_pscore = score
            best_pi = i
    if best_pi >= 0:
        m["program"] = best_pi

    return m


def _apps_parse_headers_from_table(table_root: WebElement, prefer_aria_colindex: bool = True) -> List[Tuple[int, str]]:
    """
    Build (logical_col_index, normalized_header) pairs.

    For React grids, headers often need aria-colindex to align with row cells; DOM order alone can be wrong.
    """
    headers: List[Tuple[int, str]] = []
    # Classic <table>
    try:
        ths = table_root.find_elements(By.CSS_SELECTOR, "thead th")
        if ths:
            for i, th in enumerate(ths):
                txt = (th.text or "").strip()
                if not txt:
                    continue
                col = th.get_attribute("aria-colindex")
                if prefer_aria_colindex and col and str(col).isdigit():
                    headers.append((int(col) - 1, _norm_header(txt)))
                else:
                    headers.append((i, _norm_header(txt)))
            if headers:
                return headers
    except Exception:
        pass

    # React grid headers
    try:
        hs = table_root.find_elements(By.CSS_SELECTOR, "[role='columnheader']")
        out: List[Tuple[int, str]] = []
        seq = 0
        for h in hs:
            txt = (h.text or "").strip()
            if not txt:
                continue
            col = h.get_attribute("aria-colindex")
            if prefer_aria_colindex and col and str(col).isdigit():
                out.append((int(col) - 1, _norm_header(txt)))
            else:
                out.append((seq, _norm_header(txt)))
                seq += 1
        if out:
            return out
    except Exception:
        pass

    return headers


def _apps_cells_by_col_index(row: WebElement) -> Tuple[Dict[int, WebElement], List[WebElement]]:
    """
    Return:
    - map of 0-based logical column index -> cell element (when aria-colindex is present)
    - fallback ordered cell list (td/gridcells)
    """
    cells: List[WebElement] = _row_cells(row)
    by_col: Dict[int, WebElement] = {}
    used_cols = 0
    for c in cells:
        try:
            col = c.get_attribute("aria-colindex")
            if col and str(col).isdigit():
                by_col[int(col) - 1] = c
                used_cols += 1
        except Exception:
            continue

    # If most cells don't expose colindex, treat mapping as unreliable and rely on list indexing.
    if cells and used_cols < max(2, int(0.5 * len(cells))):
        return {}, cells
    return by_col, cells


def _apps_cell_get(
    by_col: Dict[int, WebElement],
    cells: List[WebElement],
    idx: int,
) -> Optional[WebElement]:
    if idx < 0:
        return None
    if by_col:
        return by_col.get(idx)
    if 0 <= idx < len(cells):
        return cells[idx]
    return None


def get_students_header_map(driver: webdriver.Chrome) -> Dict[str, int]:
    """
    Build a column->index map by reading the visible header texts.
    This avoids wrong extraction when columns shift or when a grid adds extra columns.
    """
    headers: List[Tuple[int, str]] = []

    # Try table headers first
    try:
        ths = driver.find_elements(By.CSS_SELECTOR, "table thead th")
        if ths:
            for i, t in enumerate(ths):
                headers.append((i, _norm_header(t.text)))
    except Exception:
        pass

    # React grid fallback: use aria-colindex when available (most robust)
    if not headers:
        try:
            hs = driver.find_elements(By.CSS_SELECTOR, "[role='columnheader']")
            for h in hs:
                txt = (h.text or "").strip()
                if not txt:
                    continue
                col = h.get_attribute("aria-colindex")
                if col and str(col).isdigit():
                    headers.append((int(col) - 1, _norm_header(txt)))
                else:
                    headers.append((len(headers), _norm_header(txt)))
        except Exception:
            headers = []

    # Map known columns
    mapping: Dict[str, int] = {}
    for i, h in headers:
        if not h:
            continue
        if "student id" in h:
            mapping["student_id"] = i
        elif "student email" in h:
            mapping["student_email"] = i
        elif h == "first name" or "first name" in h:
            mapping["first_name"] = i
        elif h == "last name" or "last name" in h:
            mapping["last_name"] = i
        elif "nationality" in h:
            mapping["nationality"] = i
        elif "recruitment par" in h or "recruitment partner" in h:
            mapping["recruitment_partner"] = i
        elif "recruiter type" in h:
            mapping["recruiter_type"] = i
        elif "education" in h:
            mapping["education"] = i
    return mapping


def parse_row_by_headers(cells: List[WebElement], header_map: Dict[str, int]) -> Dict[str, str]:
    # Prefer aria-colindex if cells expose it (grid layouts)
    cell_by_col: Dict[int, WebElement] = {}
    try:
        for c in cells:
            col = c.get_attribute("aria-colindex")
            if col and str(col).isdigit():
                cell_by_col[int(col) - 1] = c
    except Exception:
        pass

    def get(key: str) -> str:
        idx = header_map.get(key, -1)
        if idx < 0:
            return ""
        if idx in cell_by_col:
            return safe_text(cell_by_col[idx])
        if 0 <= idx < len(cells):
            return safe_text(cells[idx])
        return ""

    return {
        "student_id": get("student_id"),
        "student_email": get("student_email"),
        "first_name": get("first_name"),
        "last_name": get("last_name"),
        "nationality": get("nationality"),
        "recruitment_partner": get("recruitment_partner"),
        "recruiter_type": get("recruiter_type"),
        "education": get("education"),
    }


def scrape_applications_tab(driver: webdriver.Chrome, wait: WebDriverWait, err_dir: Path, student_id: str) -> Dict[str, str]:
    """
    Extract application details from BOTH:
    - Applications (paid)
    - Unpaid Applications (unpaid/cart)

    Fills Excel columns:
    - Target University (School)
    - Target Program (Program)
    - Target Intake (Intake dropdown value if present)
    - Status App (if the paid table has Status column)
    """
    base = {
        "target_university": "",
        "target_program": "",
        "target_intake": "",
        "status_app": "",
        "destination_country": "",
        "city": "",
        "intake_label": "",
    }

    def _dump_apps_debug(tag: str) -> None:
        """
        Save artifacts to help debug DOM differences.
        Only called when we see evidence of rows but extraction is empty.
        """
        try:
            safe_sid = re.sub(r"[^a-zA-Z0-9_-]+", "_", student_id or "unknown")
            # Always try screenshot first (cheap + highest signal)
            try:
                screenshot_on_error(driver, err_dir, f"apps_{safe_sid}_{tag}")
            except Exception:
                pass
            try:
                html_path = err_dir / f"{now_stamp()}_apps_{safe_sid}_{tag}.html"
                html_path.write_text(driver.page_source or "", encoding="utf-8", errors="ignore")
            except Exception:
                pass
        except Exception:
            pass

    def _debug_table_root(table_root: WebElement, tag: str) -> None:
        try:
            hs = _table_header_texts(table_root)
        except Exception:
            hs = []
        try:
            tr = table_root.find_elements(By.CSS_SELECTOR, "tbody tr")
        except Exception:
            tr = []
        try:
            rr = table_root.find_elements(By.CSS_SELECTOR, "[role='row']")
        except Exception:
            rr = []

        sample = ""
        try:
            # Prefer a row with text
            for r in tr + rr:
                txt = (r.text or "").strip()
                if txt:
                    sample = txt.replace("\n", " ")[:180]
                    break
        except Exception:
            pass

        try:
            print(
                f"[DBG] {student_id} {tag}: url={driver.current_url} headers={hs[:8]} tbody_tr={len(tr)} role_row={len(rr)} sample={sample or '-'}",
                flush=True,
            )
        except Exception:
            pass

    def _student_base_url() -> Optional[str]:
        u = driver.current_url or ""
        m = re.search(r"(https?://[^/]+/students/\d+)", u)
        if m:
            return m.group(1)
        # SPA / redirects: recover from profile links in the DOM
        try:
            for a in driver.find_elements(By.CSS_SELECTOR, "a[href*='/students/']"):
                href = (a.get_attribute("href") or "").strip()
                m2 = re.search(r"(https?://[^/]+/students/\d+)", href)
                if m2:
                    return m2.group(1)
        except Exception:
            pass
        return None

    def _navigate_student_section(section_slug: str) -> bool:
        """
        Direct navigation is the most reliable way to avoid accidentally clicking sidebar links.
        Examples:
          .../students/2330563/unpaid-applications
          .../students/2330563/applications
        """
        base = _student_base_url()
        if not base:
            return False
        target = f"{base}/{section_slug}"
        try:
            if (driver.current_url or "").rstrip("/") == target.rstrip("/"):
                return True
            driver.get(target)
            wait_for_page(driver)
            time.sleep(0.6)
            return True
        except Exception:
            return False

    def _find_profile_tab_container() -> Optional[WebElement]:
        """
        Find the top profile tabs container (NOT the sidebar):
        Profile | Search and Apply | Applications | Unpaid Applications
        """
        try:
            # Tablist pattern (common in React UIs)
            tablists = driver.find_elements(By.XPATH, "//*[@role='tablist']")
            for tl in tablists:
                try:
                    txt = (tl.text or "").lower()
                    if (
                        "profile" in txt
                        and "applications" in txt
                        and "search and apply" in txt
                        and "unpaid applications" in txt
                    ):
                        return tl
                except Exception:
                    continue

            # XPath container that includes key labels (avoid sidebar by requiring multiple labels)
            candidates = driver.find_elements(
                By.XPATH,
                "//*[contains(., 'Profile') and contains(., 'Search and Apply') and contains(., 'Unpaid Applications') and contains(., 'Applications')]",
            )
            for c in candidates:
                try:
                    if not c.is_displayed():
                        continue
                except Exception:
                    continue
                txt = (c.text or "").lower()
                if ("profile" in txt) and ("search and apply" in txt) and ("unpaid applications" in txt):
                    return c
        except Exception:
            pass
        return None

    def click_profile_tab(tab_text: str) -> bool:
        """
        IMPORTANT: Prefer clicking the real in-profile tabs (top bar) for each student.
        Only fall back to direct URL navigation if the click cannot be performed reliably.
        """

        def _click_and_wait(tab_el: WebElement) -> bool:
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center', inline:'center'});",
                    tab_el,
                )
                time.sleep(0.2)
            except Exception:
                pass

            # Remember current state so we can detect a change.
            before_url = driver.current_url or ""
            before_hash = ""
            try:
                before_hash = (driver.find_element(By.TAG_NAME, "body").text or "")[:500]
            except Exception:
                pass

            try:
                click_with_retry(driver, tab_el, timeout_s=12, retries=3)
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", tab_el)
                except Exception:
                    return False

            # Wait until tab becomes selected OR the URL/content changes.
            def changed(d: webdriver.Chrome) -> bool:
                try:
                    # aria-selected is the most explicit signal
                    try:
                        if (tab_el.get_attribute("aria-selected") or "").lower() == "true":
                            return True
                    except Exception:
                        pass

                    cur = d.current_url or ""
                    if cur and cur != before_url:
                        return True

                    try:
                        now_hash = (d.find_element(By.TAG_NAME, "body").text or "")[:500]
                        if now_hash and now_hash != before_hash:
                            return True
                    except Exception:
                        pass
                except Exception:
                    return False
                return False

            try:
                WebDriverWait(driver, 12).until(changed)
            except Exception:
                # Sometimes tab UI changes without URL changes; still allow.
                pass

            time.sleep(0.6)
            return True

        # 1) Click the real tab in the profile tab bar (NOT sidebar)
        try:
            container = _find_profile_tab_container()
            if container is not None:
                candidates = container.find_elements(
                    By.XPATH,
                    f".//*[@role='tab' or self::a or self::button]"
                    f"[normalize-space()='{tab_text}' or .//span[normalize-space()='{tab_text}']]",
                )
                if not candidates:
                    candidates = container.find_elements(
                        By.XPATH,
                        f".//*[@role='tab' or self::a or self::button][contains(normalize-space(.), '{tab_text}')]",
                    )
                for tab in candidates[:3]:
                    try:
                        if tab.is_displayed() and tab.is_enabled() and _click_and_wait(tab):
                            return True
                    except Exception:
                        continue
        except Exception:
            pass

        # 2) Fallback: direct navigation (still in-profile, avoids sidebar)
        if tab_text == "Applications":
            return _navigate_student_section("applications")
        if tab_text == "Unpaid Applications":
            return _navigate_student_section("unpaid-applications")
        return False

    def _intake_placeholder(s: str) -> bool:
        t = (s or "").strip().lower()
        if not t:
            return True
        return t in ("select", "choose", "pick", "loading", "n/a", "na", "tbd", "--", "-")

    def _intake_from_select(sel: WebElement) -> str:
        def _skip(s: str) -> bool:
            return _intake_placeholder(s or "")

        val = (sel.get_attribute("value") or "").strip()
        if val and not _skip(val):
            return val
        try:
            opt = sel.find_element(By.CSS_SELECTOR, "option:checked")
            ot = (opt.text or "").strip()
            ov = (opt.get_attribute("value") or "").strip()
            if ot and not _skip(ot):
                return ot
            if ov and not _skip(ov):
                return ov
        except Exception:
            pass
        try:
            for opt in sel.find_elements(By.CSS_SELECTOR, "option"):
                if (opt.get_attribute("selected") or "").lower() == "true":
                    ot = (opt.text or "").strip()
                    ov = (opt.get_attribute("value") or "").strip()
                    if ot and not _skip(ot):
                        return ot
                    if ov and not _skip(ov):
                        return ov
        except Exception:
            pass
        try:
            fo = Select(sel).first_selected_option
            ot = (fo.text or "").strip()
            ov = (fo.get_attribute("value") or "").strip()
            if ot and not _skip(ot):
                return ot
            if ov and not _skip(ov):
                return ov
        except Exception:
            pass
        return ""

    def _intake_date_from_cell_text(txt: str) -> str:
        t = (txt or "").replace("\n", " ")
        t = re.sub(r"\s*Loading\s*$", "", t, flags=re.I).strip()
        m = re.search(
            r"\b(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s+\d{1,2},\s*\d{4}\b",
            t,
            re.I,
        )
        if m:
            return m.group(0)
        m = re.search(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4}\b", t, re.I)
        if m:
            return m.group(0)
        return ""

    def pick_intake_from_cell(cell: Optional[WebElement]) -> str:
        if cell is None:
            return ""
        try:
            sel = cell.find_element(By.CSS_SELECTOR, "select")
            got = _intake_from_select(sel)
            if got:
                return got
        except Exception:
            pass
        txt = (cell.text or "").replace("\n", " ")
        txt = re.sub(r"\s*Loading\s*$", "", txt, flags=re.I).strip()
        if txt and not _intake_placeholder(txt):
            dt_guess = _intake_date_from_cell_text(txt)
            return dt_guess or txt.strip()
        return _intake_date_from_cell_text(txt)

    def pick_intake_from_row(row: WebElement) -> str:
        try:
            sel = row.find_element(By.CSS_SELECTOR, "select")
            got = _intake_from_select(sel)
            if got:
                return got
        except Exception:
            pass
        try:
            txt = (row.text or "").replace("\n", " ")
            txt = re.sub(r"\s*Loading\s*$", "", txt, flags=re.I)
            m = re.search(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4}\b", txt, re.I)
            if m:
                return m.group(0)
        except Exception:
            pass
        return ""

    def _brief_wait_intake_ready(cell: Optional[WebElement]) -> None:
        if cell is None:
            return
        for _ in range(10):
            t = safe_text(cell)
            if not _intake_placeholder(t):
                return
            try:
                sel = cell.find_element(By.CSS_SELECTOR, "select")
                if _intake_from_select(sel):
                    return
            except Exception:
                pass
            time.sleep(0.2)

    def _table_header_texts(table: WebElement) -> List[str]:
        try:
            ths = table.find_elements(By.CSS_SELECTOR, "thead th")
            if ths:
                return [_norm_header(th.text) for th in ths if (th.text or "").strip()]
        except Exception:
            pass
        # React grid/table fallback
        try:
            hs = table.find_elements(By.CSS_SELECTOR, "[role='columnheader']")
            out = []
            for h in hs:
                t = (h.text or "").strip()
                if t:
                    out.append(_norm_header(t))
            if out:
                return out
        except Exception:
            pass
        return []

    def _candidate_table_roots(scope: Optional[WebElement] = None) -> List[WebElement]:
        roots: List[WebElement] = []
        search_root = scope if scope is not None else driver
        try:
            roots.extend(search_root.find_elements(By.CSS_SELECTOR, "table"))
        except Exception:
            pass
        # React grid variants (some pages don’t use <table>)
        try:
            roots.extend(search_root.find_elements(By.CSS_SELECTOR, "[role='table'], [role='grid']"))
        except Exception:
            pass
        # Deduplicate by id in memory
        seen: set[int] = set()
        out: List[WebElement] = []
        for r in roots:
            try:
                k = id(r)
                if k in seen:
                    continue
                seen.add(k)
                out.append(r)
            except Exception:
                continue
        return out

    def _pick_best_table(required: List[str], scope: Optional[WebElement] = None) -> Optional[WebElement]:
        """
        Find the most likely applications table in the active tab content by scoring headers.
        """
        best: Optional[WebElement] = None
        best_score = 0
        for t in _candidate_table_roots(scope):
            try:
                if not t.is_displayed():
                    continue
            except Exception:
                pass
            # Ignore tiny tables / unrelated chrome tables
            try:
                if t.size.get("height", 9999) < 40:
                    continue
            except Exception:
                pass
            hs = _table_header_texts(t)
            if not hs:
                continue
            score = 0
            for r in required:
                if any(r in h for h in hs):
                    score += 2
            if any("intake" in h for h in hs):
                score += 2
            if any("school" in h or "university" in h or "institution" in h for h in hs):
                score += 2
            if any("program" in h for h in hs):
                score += 2
            if score > best_score:
                best_score = score
                best = t
        return best

    def _find_section_root(title: str) -> Optional[WebElement]:
        """
        Prefer extracting tables from the correct section (Paid vs Unpaid) to avoid picking unrelated grids.
        """
        safe = (title or "").replace("\"", "")
        if not safe:
            return None
        try:
            secs = driver.find_elements(
                By.XPATH,
                "//*[self::section or self::div][.//*[contains(normalize-space(.), \"" + safe + "\")]]",
            )
        except Exception:
            return None
        for s in secs:
            try:
                if s.is_displayed():
                    return s
            except Exception:
                continue
        return None

    def _wait_tab_table_ready(tab_name: str, required_keywords: Optional[List[str]] = None, timeout_s: int = 30) -> None:
        """
        After switching tabs, the table often renders asynchronously.
        Wait for either:
        - a visible table with >=1 tbody row, OR
        - a visible table/grid with column headers, OR
        - a visible empty-state message.
        Never throws (best-effort).
        """

        def ready(d: webdriver.Chrome) -> bool:
            try:
                # Any visible table with real rows?
                roots = _candidate_table_roots()
                for t in roots:
                    try:
                        if not t.is_displayed():
                            continue
                    except Exception:
                        continue
                    try:
                        # Avoid header-only tables
                        rows = t.find_elements(By.CSS_SELECTOR, "tbody tr")
                        row_texts = [(r.text or "").strip() for r in rows if (r.text or "").strip()]
                        if row_texts:
                            # Wait out transient placeholders like "Loading"
                            if all(rt.lower() == "loading" for rt in row_texts):
                                return False
                            if any("loading" in rt.lower() for rt in row_texts) and not any(
                                re.search(r"\b\d{6,}\b", rt) for rt in row_texts
                            ) and not any("view" in rt.lower() for rt in row_texts):
                                return False
                            # Empty-state row is acceptable
                            elif any("no unpaid applications" in rt.lower() or "no applications" in rt.lower() for rt in row_texts):
                                return True
                            # Real data row (application id present)
                            elif any(re.search(r"\b\d{6,}\b", rt) for rt in row_texts):
                                return True
                            # Another strong signal of real rows: View button + a month/year intake
                            elif any("view" in rt.lower() for rt in row_texts) and any(
                                re.search(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4}\b", rt, re.I)
                                for rt in row_texts
                            ):
                                return True
                            else:
                                # Some tables may not show numeric ids; treat any non-loading text as ready
                                return True
                    except Exception:
                        pass
                    # React grid rows
                    try:
                        rows = t.find_elements(By.CSS_SELECTOR, "[role='row']")
                        # Many grids include header row(s); require at least a couple of non-empty rows.
                        row_texts = [(r.text or "").strip() for r in rows if (r.text or "").strip()]
                        if len(row_texts) >= 2:
                            if any("loading" in rt.lower() for rt in row_texts) and not any(
                                re.search(r"\b\d{6,}\b", rt) for rt in row_texts
                            ):
                                pass
                            else:
                                return True
                    except Exception:
                        pass
                    # Column headers rendered
                    try:
                        if len(t.find_elements(By.CSS_SELECTOR, "thead th")) >= 3:
                            return True
                    except Exception:
                        pass
                    try:
                        if len(t.find_elements(By.CSS_SELECTOR, "[role='columnheader']")) >= 3:
                            return True
                    except Exception:
                        pass

                # Empty states / no-data messages
                body_txt = (d.find_element(By.TAG_NAME, "body").text or "").lower()
                if "no applications" in body_txt or "no records" in body_txt or "no results" in body_txt:
                    return True

                # If we know which keywords should exist (e.g., School/Program/Intake), wait for them.
                if required_keywords:
                    hits = 0
                    for kw in required_keywords:
                        if (kw or "").strip().lower() in body_txt:
                            hits += 1
                    if hits >= max(1, min(2, len(required_keywords))):
                        return True
            except Exception:
                return False
            return False

        try:
            WebDriverWait(driver, timeout_s).until(ready)
        except Exception:
            pass

    def extract_all_rows_from_table(table_root: WebElement) -> Dict[str, List[str]]:
        out: Dict[str, List[str]] = {
            "target_university": [],
            "target_program": [],
            "target_intake": [],
            "status_app": [],
        }

        # Row extraction
        try:
            rows = table_root.find_elements(By.CSS_SELECTOR, "tbody tr")
            if not rows:
                rows = table_root.find_elements(By.CSS_SELECTOR, "[role='row']")

            prefer_aria = True
            for probe in rows[:10]:
                by_c, cs = _apps_cells_by_col_index(probe)
                if len(cs) < 3:
                    continue
                prefer_aria = bool(by_c and len(by_c) >= max(2, int(0.5 * len(cs))))
                break

            headers_list = _apps_parse_headers_from_table(table_root, prefer_aria_colindex=prefer_aria)
            base_m: Dict[str, int] = _apps_table_column_map(headers_list)

            for r in rows:
                m = dict(base_m)
                txt = (r.text or "").strip()
                if not txt:
                    continue
                by_col, cells = _apps_cells_by_col_index(r)
                if len(cells) < 3:
                    continue
                try:
                    if len(r.find_elements(By.CSS_SELECTOR, "input, select, textarea")) >= 4:
                        continue
                except Exception:
                    pass

                def cell_raw(i: int) -> str:
                    el = _apps_cell_get(by_col, cells, i)
                    return safe_text(el) if el is not None else ""

                # Typical unpaid/cart grid: application Id (6+ digits) | School | Program | Intake | ...
                # Do NOT infer when col0 is Priority (1–2 digits) or other small ids.
                if "school" not in m and len(cells) >= 4:
                    c0 = (cell_raw(0) or "").strip()
                    if re.fullmatch(r"\d{6,}", c0):
                        m.setdefault("school", 1)
                        m.setdefault("program", 2)
                        m.setdefault("intake", 3)

                si = m.get("school", 1)
                pi = m.get("program", 2)
                req_i = m.get("requirements", -1)

                school_el = _apps_cell_get(by_col, cells, si)
                prog_el = _apps_cell_get(by_col, cells, pi)
                # University/School: always from the School column (visible text or logo alt inside that cell).
                school = _text_or_img_alt(school_el) if school_el is not None else ""
                program = _program_cell_text(prog_el) if prog_el is not None else ""
                if not program and prog_el is not None:
                    program = _text_or_img_alt(prog_el)

                st_i = m.get("status", -1)
                status = cell_raw(st_i) if isinstance(st_i, int) and st_i >= 0 else ""
                status_el = _apps_cell_get(by_col, cells, st_i) if isinstance(st_i, int) and st_i >= 0 else None
                if status_el is not None:
                    try:
                        for el in status_el.find_elements(By.CSS_SELECTOR, "strong, b"):
                            t = (safe_text(el) or "").strip()
                            if t and len(t) < 120 and not _is_applyboard_substatus_line(t):
                                status = t
                                break
                    except Exception:
                        pass

                # If mapped Program column is actually Requirements / workflow text, pick real program from row.
                if (not program) or _is_application_stage_or_req_token(program):
                    best_prog = ""
                    col_cells: List[Tuple[int, WebElement]] = []
                    if by_col:
                        for k, c in sorted(by_col.items(), key=lambda t: t[0]):
                            col_cells.append((k, c))
                    else:
                        for i, c in enumerate(cells):
                            col_cells.append((i, c))
                    for idx, c in col_cells:
                        if idx == si:
                            continue
                        if isinstance(req_i, int) and req_i >= 0 and idx == req_i:
                            continue
                        cand = (_program_cell_text(c) or _text_or_img_alt(c) or "").strip()
                        if not cand:
                            continue
                        if _is_application_stage_or_req_token(cand):
                            continue
                        if not _looks_like_program_name(cand):
                            continue
                        if len(cand) > len(best_prog):
                            best_prog = cand
                    if best_prog:
                        program = best_prog

                # If "program" is actually a workflow label, treat it as Status App instead.
                if program and _is_application_stage_or_req_token(program) and not _looks_like_program_name(program):
                    if not status:
                        status = program
                    program = ""

                # School column should only contain institution text + logos; if we still see program-like text there,
                # fall back to logo alt inside the same cell only (do not pull from other columns).
                if _looks_like_program_name(school) and school_el is not None:
                    fixed = ""
                    try:
                        for img in school_el.find_elements(By.CSS_SELECTOR, "img[alt]"):
                            alt = (img.get_attribute("alt") or "").strip()
                            if len(alt) > 3 and not _looks_like_program_name(alt):
                                fixed = alt
                                break
                    except Exception:
                        pass
                    if fixed:
                        if not program:
                            program = school
                        school = fixed

                if not school and program and _looks_like_program_name(program) and school_el is not None:
                    try:
                        for img in school_el.find_elements(By.CSS_SELECTOR, "img[alt]"):
                            alt = (img.get_attribute("alt") or "").strip()
                            if len(alt) > 3 and not _looks_like_program_name(alt):
                                school = alt
                                break
                    except Exception:
                        pass

                if not school and not program:
                    continue

                # Clean common UI helper text
                for bad in ["(Opens in new tab)", "Opens in new tab"]:
                    school = school.replace(bad, "").strip()
                    program = program.replace(bad, "").strip()

                in_i = m.get("intake", -1)
                in_el = _apps_cell_get(by_col, cells, in_i) if isinstance(in_i, int) and in_i >= 0 else None
                _brief_wait_intake_ready(in_el)
                raw_intake = safe_text(in_el) if in_el is not None else ""
                picked = pick_intake_from_cell(in_el) or pick_intake_from_row(r)
                if _intake_placeholder(raw_intake):
                    intake = picked
                else:
                    intake = raw_intake.strip()
                if _intake_placeholder(intake):
                    intake = ""

                status = _primary_paid_application_status((status or "").strip())
                out["target_university"].append(school)
                out["target_program"].append(program)
                out["status_app"].append(status)
                out["target_intake"].append(intake)
        except Exception:
            return out

        return out

    def _wait_paid_apps_loaded(timeout_s: int = 45) -> None:
        """
        Paid Applications page often shows a spinner before rows are rendered.
        Wait for spinner to disappear OR for a visible 'View' button in Paid Applications section.
        Best-effort; never raises.
        """

        def ready(d: webdriver.Chrome) -> bool:
            try:
                # If spinner is visible, we're not ready yet.
                spinners = d.find_elements(By.CSS_SELECTOR, ".react-spinners--clip")
                for s in spinners:
                    try:
                        if s.is_displayed():
                            return False
                    except Exception:
                        continue

                # Look for 'Paid Applications' section and a 'View' button
                try:
                    paid_sections = d.find_elements(
                        By.XPATH,
                        "//section[.//*[contains(normalize-space(.), 'Paid Applications')]]",
                    )
                except Exception:
                    paid_sections = []

                for sec in paid_sections:
                    try:
                        if not sec.is_displayed():
                            continue
                    except Exception:
                        pass
                    sec_txt = ""
                    try:
                        sec_txt = (sec.text or "").strip()
                    except Exception:
                        sec_txt = ""
                    try:
                        views = sec.find_elements(
                            By.XPATH,
                            ".//*[self::a or self::button][contains(normalize-space(.), 'View')]",
                        )
                        if any(v.is_displayed() for v in views):
                            return True
                    except Exception:
                        pass

                    # Real data rows often include an App # (5+ digits) and/or a status like Processing
                    if sec_txt:
                        low = sec_txt.lower()
                        if "loading" in low:
                            continue
                        if re.search(r"\b\d{5,}\b", sec_txt):
                            return True
                        if any(
                            x in low
                            for x in (
                                "processing",
                                "rejected",
                                "accepted",
                                "canceled",
                                "cancelled",
                                "submitted",
                                "draft",
                            )
                        ):
                            return True

                    # Empty state: no paid applications text
                    try:
                        txt = (sec_txt or "").lower()
                        if "no paid applications" in txt or "no applications" in txt:
                            return True
                    except Exception:
                        pass

            except Exception:
                return False
            return False

        try:
            WebDriverWait(driver, timeout_s).until(ready)
        except Exception:
            pass

    def _extract_paid_from_view_rows() -> Dict[str, List[str]]:
        """
        Paid Applications sometimes isn't a <table>. Use 'View' rows inside the 'Paid Applications' section.
        Heuristic extraction of School + Program + (Start Date as intake-like value if present).
        """
        out: Dict[str, List[str]] = {"target_university": [], "target_program": [], "target_intake": [], "status_app": []}
        try:
            secs = driver.find_elements(By.XPATH, "//section[.//*[contains(normalize-space(.), 'Paid Applications')]]")
        except Exception:
            return out

        header_stop = {
            "priority",
            "status",
            "app #",
            "app#",
            "school",
            "program",
            "esl start date",
            "start date",
            "requirements",
            "paid applications",
            "looking for your unpaid applications?",
        }

        for sec in secs:
            try:
                if not sec.is_displayed():
                    continue
            except Exception:
                pass

            try:
                view_buttons = sec.find_elements(
                    By.XPATH, ".//*[self::a or self::button][contains(normalize-space(.), 'View')]"
                )
            except Exception:
                view_buttons = []

            for vb in view_buttons:
                try:
                    if not vb.is_displayed():
                        continue
                except Exception:
                    pass

                # Climb up to a row-ish container that holds the text for this application.
                row_el = vb
                for _ in range(8):
                    try:
                        row_el = row_el.find_element(By.XPATH, "..")
                    except Exception:
                        break
                    txt = (row_el.text or "").strip()
                    if txt and "paid applications" not in txt.lower() and "priority" not in txt.lower():
                        low = txt.lower()
                        # Row container: app # and/or headline application status
                        if re.search(r"\b\d{5,}\b", txt) or any(
                            k in low
                            for k in (
                                "processing",
                                "rejected",
                                "accepted",
                                "canceled",
                                "cancelled",
                                "submitted",
                                "draft",
                                "admitted",
                                "deferred",
                            )
                        ):
                            break

                txt = (row_el.text or "").strip()
                if not txt:
                    continue

                lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
                # Remove UI helper/header labels
                cleaned: List[str] = []
                for ln in lines:
                    low = ln.lower()
                    if low in header_stop:
                        continue
                    if low == "view":
                        continue
                    cleaned.append(ln.replace("(Opens in new tab)", "").strip())

                app_idx = None
                for i, ln in enumerate(cleaned):
                    if re.fullmatch(r"\d{5,}", ln):
                        app_idx = i
                        break
                if app_idx is None:
                    # Sometimes app # is attached to other text; search each line.
                    for i, ln in enumerate(cleaned):
                        if re.search(r"\b\d{5,}\b", ln):
                            app_idx = i
                            break

                # Heuristic: after app number comes School then Program
                school = ""
                program = ""

                # Paid Applications often shows School as a logo; prefer img alt first (e.g. "Niagara College").
                try:
                    for img in row_el.find_elements(By.CSS_SELECTOR, "img[alt]"):
                        alt = (img.get_attribute("alt") or "").strip()
                        if len(alt) < 4 or _looks_like_program_name(alt):
                            continue
                        if re.search(r"\b(college|university|institute|polytechnic|academy|campus)\b", alt, re.I):
                            school = alt
                            break
                except Exception:
                    pass

                # Prefer explicit school names if present — but DO NOT treat program titles like
                # "College Diploma - ..." as the school just because they contain the word "college".
                if not school:
                    for ln in cleaned:
                        if not re.search(r"\b(college|university|institute|polytechnic|academy|campus)\b", ln, re.I):
                            continue
                        if _looks_like_program_name(ln):
                            continue
                        school = ln
                        break

                # Program is usually the most "descriptive" line (often contains '-' or a degree keyword).
                prog_candidates: List[str] = []
                for ln in cleaned:
                    low = ln.lower()
                    if low in ("processing", "submitted", "draft"):
                        continue
                    if re.fullmatch(r"\d{5,}", ln):
                        continue
                    if ln == school:
                        continue
                    if re.search(r"\b\d{4}\s*[-/]\s*[A-Za-z]{3,}\b", ln):
                        continue
                    if low in ("esl", "n/a", "na"):
                        continue
                    if len(ln) < 8:
                        continue
                    if "-" in ln or re.search(r"\b(master|bachelor|diploma|certificate|program)\b", ln, re.I):
                        prog_candidates.append(ln)

                if prog_candidates:
                    # choose the longest descriptive candidate
                    program = sorted(prog_candidates, key=lambda s: len(s), reverse=True)[0]

                # If we still have nothing, try simple positional fallback after app number
                if not program and app_idx is not None:
                    after = cleaned[app_idx + 1 :]
                    if after:
                        program = after[0]

                # School often renders as logo-only: use img alt (e.g. "Northeastern University")
                if not school:
                    try:
                        for img in row_el.find_elements(By.CSS_SELECTOR, "img[alt]"):
                            alt = (img.get_attribute("alt") or "").strip()
                            if len(alt) > 3 and not _looks_like_program_name(alt):
                                school = alt
                                break
                    except Exception:
                        pass

                intake_like = ""
                # Paid page often shows Start Date like "2026-Sep"
                for ln in cleaned:
                    m = re.search(r"\b\d{4}\s*[-/]\s*[A-Za-z]{3,}\b", ln)
                    if m:
                        intake_like = m.group(0).replace(" ", "")
                        break

                status = _primary_paid_application_status("\n".join(cleaned))

                if school or program:
                    out["target_university"].append(school)
                    out["target_program"].append(program)
                    out["target_intake"].append(intake_like or "")
                    out["status_app"].append(status or "")

        return out

    # Paid Applications (inside Applications tab)
    paid_rows: Dict[str, List[str]] = {"target_university": [], "target_program": [], "target_intake": [], "status_app": []}
    if click_profile_tab("Applications"):
        _wait_paid_apps_loaded()
        _wait_tab_table_ready("Applications", required_keywords=["school", "program"])
        paid_scope = _find_section_root("Paid Applications") or _find_section_root("Applications")
        t = _pick_best_table(required=["school", "program"], scope=paid_scope)
        paid_from_table = False
        if t is not None:
            _debug_table_root(t, "paid_table")
            paid_rows = extract_all_rows_from_table(t)
            paid_from_table = True
        else:
            # Could not identify the table root for paid applications; dump DOM for analysis.
            _dump_apps_debug("paid_no_table")
            paid_rows = _extract_paid_from_view_rows()

        if not paid_rows["target_university"] and not paid_rows["target_program"]:
            # If the UI has rows but we extracted nothing, dump artifacts.
            try:
                if len(driver.find_elements(By.CSS_SELECTOR, "table tbody tr")) >= 1 or len(driver.find_elements(By.CSS_SELECTOR, "[role='row']")) >= 3:
                    _dump_apps_debug("paid_empty")
            except Exception:
                pass
        elif paid_from_table:
            # Table parsing can grab the wrong grid or mis-map columns; fall back to Paid Applications row heuristics.
            fb = _extract_paid_from_view_rows()
            if fb["target_university"] or fb["target_program"]:

                def _univ_looks_plausible(u: str) -> bool:
                    t = (u or "").strip().lower()
                    if len(t) < 6:
                        return False
                    if _looks_like_program_name(u) and not re.search(
                        r"\b(college|university|institute|polytechnic|academy)\b", t
                    ):
                        return False
                    return bool(re.search(r"\b(college|university|institute|polytechnic|academy|campus)\b", t))

                def _score_univ_list(us: List[str]) -> int:
                    return sum(1 for x in (us or []) if _univ_looks_plausible(x))

                if _score_univ_list(fb["target_university"]) > _score_univ_list(paid_rows["target_university"]):
                    paid_rows = fb

    # Unpaid Applications tab
    unpaid_rows: Dict[str, List[str]] = {"target_university": [], "target_program": [], "target_intake": [], "status_app": []}
    if click_profile_tab("Unpaid Applications"):
        _wait_tab_table_ready("Unpaid Applications", required_keywords=["school", "program", "intake"])
        # This tab clearly has Id/School/Program/Intake columns (per your screenshot)
        unpaid_scope = _find_section_root("Unpaid Applications")
        t = _pick_best_table(required=["school", "program", "intake"], scope=unpaid_scope)
        if t is not None:
            _debug_table_root(t, "unpaid_table")
            unpaid_rows = extract_all_rows_from_table(t)
        if not unpaid_rows["target_university"] and not unpaid_rows["target_program"]:
            try:
                if len(driver.find_elements(By.CSS_SELECTOR, "table tbody tr")) >= 1 or len(driver.find_elements(By.CSS_SELECTOR, "[role='row']")) >= 3:
                    _dump_apps_debug("unpaid_empty")
            except Exception:
                pass

    # Unpaid / cart tab does not expose the same Status column as Paid Apps — default empty to "Unpaid".
    _nu = max(
        len(unpaid_rows["target_university"]),
        len(unpaid_rows["target_program"]),
        len(unpaid_rows["target_intake"]),
        len(unpaid_rows["status_app"]),
    )
    while len(unpaid_rows["status_app"]) < _nu:
        unpaid_rows["status_app"].append("")
    for _i in range(_nu):
        if _i < len(unpaid_rows["status_app"]) and not str(unpaid_rows["status_app"][_i]).strip():
            unpaid_rows["status_app"][_i] = "Unpaid"

    # Merge paid + unpaid. Do NOT forward-fill universities across rows (that pairs the wrong school with programs).

    def _zip_aligned(us: List[str], ps: List[str], ins: List[str], sts: List[str]) -> List[Tuple[str, str, str, str]]:
        n = max(len(us), len(ps), len(ins), len(sts))

        def pad(xs: List[str]) -> List[str]:
            xs = list(xs or [])
            if len(xs) >= n:
                return xs[:n]
            return xs + [""] * (n - len(xs))

        us2, ps2, ins2, sts2 = pad(us), pad(ps), pad(ins), pad(sts)
        out: List[Tuple[str, str, str, str]] = []
        for i in range(n):
            out.append((us2[i].strip(), ps2[i].strip(), ins2[i].strip(), sts2[i].strip()))
        return out

    def _normalize_triple(u: str, p: str, inc: str, st: str) -> Tuple[str, str, str, str]:
        # If program column accidentally captured workflow tokens, treat them as status — not program.
        if _is_application_stage_or_req_token(p) and not _looks_like_program_name(p):
            if not st:
                st = p
            p = ""
        if _is_application_stage_or_req_token(u) and not _looks_like_program_name(u):
            # Rare: mis-mapped columns
            if not st:
                st = u
            u = ""
        return u, p, inc, st

    triples: List[Tuple[str, str, str, str]] = []
    triples.extend(_zip_aligned(paid_rows["target_university"], paid_rows["target_program"], paid_rows["target_intake"], paid_rows["status_app"]))
    triples.extend(_zip_aligned(unpaid_rows["target_university"], unpaid_rows["target_program"], unpaid_rows["target_intake"], unpaid_rows["status_app"]))

    seen_k: set[Tuple[str, str, str]] = set()
    merged_univ: List[str] = []
    merged_prog: List[str] = []
    merged_intake: List[str] = []
    merged_status: List[str] = []
    for u, p, inc, st in triples:
        u, p, inc, st = _normalize_triple(u, p, inc, st)
        if not u and not p:
            continue
        key = (u, p, inc)
        if key in seen_k:
            continue
        seen_k.add(key)
        merged_univ.append(u)
        merged_prog.append(p)
        merged_intake.append(inc)
        merged_status.append(st or "")

    # Destination/Country from university names (OpenAI-assisted if OPENAI_API_KEY is set)
    dests: List[str] = []
    for u in merged_univ:
        ctry = infer_destination_country_from_university(u)
        if ctry and ctry not in dests:
            dests.append(ctry)

    return {
        **base,
        "destination_country": _scrub_export_field_noise("; ".join(dests)),
        "target_university": _scrub_export_field_noise("; ".join(merged_univ)),
        "target_program": _scrub_export_field_noise("; ".join(merged_prog)),
        "target_intake": _scrub_export_field_noise("; ".join(merged_intake)),
        "status_app": _scrub_export_field_noise("; ".join(merged_status)),
    }


_STUDENT_ID_RE = re.compile(r"^\s*\d{4,}\s*$")


def _is_application_stage_or_req_token(s: str) -> bool:
    t = (s or "").strip().lower()
    if not t:
        return False
    if t in (
        "pre-submission",
        "pre submission",
        "submission",
        "draft",
        "submitted",
        "paid",
        "unpaid",
        "requirements",
        "requirement",
        "priority",
        "deferral",
        "waitlist",
        "cancelled",
        "canceled",
        "closed",
        "n/a",
        "na",
        "tbd",
    ):
        return True
    if "requirement" in t and len(t) < 40:
        return True
    return False


def _looks_like_program_name(s: str) -> bool:
    if not s or len(s) < 10:
        return False
    if _is_application_stage_or_req_token(s):
        return False
    sl = s.lower()
    if any(
        k in sl
        for k in (
            "diploma",
            "master",
            "bachelor",
            "certificate",
            "doctor",
            "mba",
            "phd",
            "graduate",
            "undergraduate",
            "technician",
            "advanced diploma",
            "optional co-op",
            "co-op",
        )
    ):
        return True
    if s.count("-") >= 1 and re.search(r"\(\d{3,5}\)", s):
        return True
    return False


def _text_or_img_alt(el: WebElement) -> str:
    """Visible text, else first meaningful img alt/title (school logos often have no text)."""
    t = safe_text(el)
    if t:
        return t
    try:
        for img in el.find_elements(By.CSS_SELECTOR, "img"):
            alt = (img.get_attribute("alt") or "").strip()
            if len(alt) > 2 and not alt.lower().startswith(("icon", "logo", "image")):
                return alt
            title = (img.get_attribute("title") or "").strip()
            if len(title) > 2:
                return title
    except Exception:
        pass
    return ""


def _program_cell_text(cell: WebElement) -> str:
    """Prefer the program link text inside the cell."""
    try:
        for a in cell.find_elements(By.CSS_SELECTOR, "a"):
            tt = safe_text(a)
            if tt and len(tt) > 6:
                return tt
    except Exception:
        pass
    return _text_or_img_alt(cell)


def _row_cells(row: WebElement) -> List[WebElement]:
    """ApplyBoard uses classic <table> rows or React grid rows (role=gridcell)."""
    try:
        tds = row.find_elements(By.TAG_NAME, "td")
        if len(tds) >= 4:
            return tds
        gcs = row.find_elements(By.CSS_SELECTOR, "[role='gridcell'], [role='cell']")
        if len(gcs) >= 4:
            return gcs
        if tds:
            return tds
        return gcs
    except Exception:
        return []


def _is_student_data_row(row: WebElement) -> bool:
    """
    ApplyBoard tables often include:
    - header row(s)
    - a filter row with inputs under column headers
    We keep rows where Student ID looks like a numeric ID.
    """
    try:
        cells = _row_cells(row)
        # Full grid can be wide (horizontal scroll); require at least Student ID + email columns.
        if len(cells) < 5:
            return False

        # Filter rows typically contain lots of <input> fields.
        try:
            inputs = row.find_elements(By.CSS_SELECTOR, "input, select, textarea")
            if len(inputs) >= 4:
                return False
        except Exception:
            pass

        sid = safe_text(cells[1])
        if not sid:
            return False
        if _STUDENT_ID_RE.match(sid):
            return True
        if re.search(r"\d{4,}", sid):
            return True

        # Fallback: Student ID column often contains a mailto/text link with digits.
        try:
            links = cells[1].find_elements(By.CSS_SELECTOR, "a")
            for a in links:
                t = (a.text or "").strip()
                if re.fullmatch(r"\d{4,}", t):
                    return True
        except Exception:
            pass

        # Loose: full row text often has "2330590 ... user@email.com" even if cell indices differ.
        try:
            row_txt = (row.text or "").replace("\n", " ")
            if re.search(r"\b\d{6,}\b", row_txt) and "@" in row_txt and len(cells) >= 3:
                return True
        except Exception:
            pass

        return False
    except Exception:
        return False


def get_table_rows(driver: webdriver.Chrome) -> List[WebElement]:
    """Prefer validated student rows; fall back to raw rows only for legacy callers."""
    dr = get_student_data_rows(driver)
    if dr:
        return dr
    for sel in TABLE_ROW_SELECTORS:
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, sel)
            if not rows:
                continue
            if sel == "table tr":
                filtered: List[WebElement] = []
                for r in rows:
                    try:
                        ths = r.find_elements(By.TAG_NAME, "th")
                        if ths:
                            continue
                    except Exception:
                        pass
                    filtered.append(r)
                if filtered:
                    rows = filtered
            return rows
        except Exception:
            continue
    return []


def get_student_data_rows(driver: webdriver.Chrome) -> List[WebElement]:
    """
    Rows that look like real students (numeric Student ID). Empty while table still shows Loading
    or only header/filter rows — use this to know the grid has finished loading.
    """
    for sel in TABLE_ROW_SELECTORS:
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, sel)
            if not rows:
                continue
            if sel == "table tr":
                filtered: List[WebElement] = []
                for r in rows:
                    try:
                        ths = r.find_elements(By.TAG_NAME, "th")
                        if ths:
                            continue
                    except Exception:
                        pass
                    filtered.append(r)
                if filtered:
                    rows = filtered

            data_rows = [r for r in rows if _is_student_data_row(r)]
            if data_rows:
                return data_rows
        except Exception:
            continue
    return []


def get_student_link_from_row(row: WebElement) -> Optional[WebElement]:
    """
    Student ID is usually column index 1, but some layouts differ — scan all cells for a profile link
    (numeric id text or /student(s)/ in href).
    """
    try:
        cells = _row_cells(row)
        if not cells:
            return None

        def score_link(a: WebElement) -> int:
            s = 0
            try:
                href = (a.get_attribute("href") or "").lower()
                if "student" in href:
                    s += 3
                if re.search(r"/\d{4,}", href):
                    s += 2
            except Exception:
                pass
            t = (a.text or "").strip()
            if re.fullmatch(r"\d{4,}", t):
                s += 3
            return s

        best: Optional[WebElement] = None
        best_score = 0
        for c in cells:
            for a in c.find_elements(By.CSS_SELECTOR, "a[href]"):
                sc = score_link(a)
                if sc > best_score:
                    best_score = sc
                    best = a
        if best is not None and best_score > 0:
            return best

        if len(cells) >= 2:
            links = cells[1].find_elements(By.CSS_SELECTOR, "a")
            for a in links:
                t = (a.text or "").strip()
                if re.fullmatch(r"\d{4,}", t):
                    return a
            if links:
                return links[0]
            return cells[1]
    except Exception:
        return None
    return None


def scrape_student_profile(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    err_dir: Path,
    student_id: str,
) -> Tuple[str, str, str, Dict[str, str]]:
    try:
        wait_for_page(driver)
        # Wait for profile content to render (not just <body>)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        WebDriverWait(driver, 45).until(
            lambda d: ("registration date" in ((d.find_element(By.TAG_NAME, "body").text or "").lower()))
            or (len(d.find_elements(By.CSS_SELECTOR, "a[href^='tel:']")) > 0)
        )
        phone = extract_phone(driver)
        reg_date = extract_registration_date(driver)
        apps = scrape_applications_tab(driver, wait, err_dir=err_dir, student_id=student_id)
        url = driver.current_url or ""
        return phone, reg_date, url, apps
    except Exception:
        screenshot_on_error(driver, err_dir, f"profile_{re.sub(r'[^a-zA-Z0-9_-]+', '_', student_id) or 'unknown'}")
        return "", "", driver.current_url or "", {}


def _open_profile_in_new_tab(driver: webdriver.Chrome, el: WebElement, student_id: str = "") -> bool:
    """
    Open a student profile in a new tab to avoid losing Students list state and avoid SPA back/refresh issues.
    Returns True if a new tab was opened.
    """
    href = ""
    try:
        href = (el.get_attribute("href") or "").strip()
    except Exception:
        pass
    if not href:
        try:
            anc = el.find_element(By.XPATH, "./ancestor-or-self::a[@href][1]")
            href = (anc.get_attribute("href") or "").strip()
        except Exception:
            pass
    sid = (student_id or "").strip()
    if not href and sid:
        href = f"{_applyboard_base_url(driver)}/students/{sid}"

    if href:
        try:
            before = list(driver.window_handles)
            driver.execute_script("window.open(arguments[0], '_blank');", href)
            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(before))
            return True
        except Exception:
            pass

    # Fallback: click in same tab (pagination may reset — caller restores footer)
    return False


def _wait_after_student_click(driver: webdriver.Chrome, before_url: str, timeout_s: int = 25) -> None:
    """
    SPA navigations often don't change hostname; wait until URL changes OR profile-ish UI appears.
    """
    def progressed(d: webdriver.Chrome) -> bool:
        try:
            cur = (d.current_url or "").lower()
            base = (before_url or "").lower()
            if cur and cur != base:
                return True
            body = (d.find_element(By.TAG_NAME, "body").text or "").lower()
            if "registration date" in body:
                return True
            return False
        except Exception:
            return False

    WebDriverWait(driver, timeout_s).until(progressed)


def _set_effective_students_url(url: str) -> None:
    global _EFFECTIVE_STUDENTS_URL
    if url:
        _EFFECTIVE_STUDENTS_URL = url


def _reload_students_list(driver: webdriver.Chrome, err_dir: Path) -> None:
    """Reload Students list via GET only (no history.back — avoids /agent)."""
    driver.get(STUDENTS_URL)
    wait_for_page(driver)
    _maybe_click_partners_splash(driver, err_dir=err_dir)
    _wait_students_list_ready(driver, err_dir=err_dir, timeout_s=180)


def _scrub_export_field_noise(s: str) -> str:
    """Remove ApplyBoard UI banners (e.g. 'This intake is closed') from scraped CSV fields."""
    if not (s or "").strip():
        return ""
    t = str(s).replace("\r", " ")
    t = re.sub(r"(?i)\s*this\s+intake\s+is\s+closed\s*", " ", t)
    t = re.sub(r"(?i)\s*intake\s+is\s+closed\s*", " ", t)
    t = re.sub(r"\s+", " ", t.replace("\n", " ")).strip()
    return t


def dismiss_applyboard_lightboxes(driver: webdriver.Chrome) -> None:
    """Close snackbars/modals that block the grid after viewing applications (prevents false empty table)."""
    try:
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium.webdriver.common.keys import Keys

        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(0.08)
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    except Exception:
        pass
    try:
        driver.execute_script(
            """
            document.querySelectorAll('[role="alert"], .MuiSnackbar-root').forEach(function(e){
              try { if (e.offsetParent !== null) e.style.display='none'; } catch(_) {}
            });
            """
        )
    except Exception:
        pass


def _applyboard_base_url(driver: webdriver.Chrome) -> str:
    try:
        u = driver.current_url or ""
        m = re.match(r"(https?://[^/]+)", u)
        if m and "applyboard.com" in m.group(1).lower():
            return m.group(1).rstrip("/")
    except Exception:
        pass
    return "https://www.applyboard.com"


def _try_click_numeric_page_button(driver: webdriver.Chrome, page_num_1based: int) -> bool:
    """
    MUI pagination often shows 1,2,…,37 — click the page number to jump in one step
    when visible (avoids 10+ slow Next clicks after a full list reload).
    """
    if page_num_1based < 1:
        return False
    txt = str(page_num_1based)
    try:
        for sel in (
            "ul.MuiPagination-ul button",
            ".MuiPagination-root button",
            "nav[aria-label*='pagination' i] button",
        ):
            for b in driver.find_elements(By.CSS_SELECTOR, sel):
                try:
                    if not b.is_displayed():
                        continue
                    if (b.text or "").strip() != txt:
                        continue
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'}); arguments[0].click();",
                        b,
                    )
                    time.sleep(0.55)
                    return True
                except Exception:
                    continue
    except Exception:
        pass
    return False


def ensure_students_list_footer_start(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    err_dir: Path,
    target_range_lo: int,
    *,
    jump_fast: bool = True,
    script_page: int = 0,
) -> None:
    """
    ApplyBoard resets the Students grid to page 1 after driver.get(/students) when returning
    from a profile in the same tab. Re-click Next until the footer window matches the row band
    we were scraping (e.g. 281 for page 15 at 20 rows/page).
    """
    if target_range_lo <= 1:
        return
    r = _read_pagination_range(driver)
    if not r:
        return
    cur_lo = r[0]
    if cur_lo == target_range_lo:
        return
    if target_range_lo <= cur_lo <= target_range_lo + 24:
        return
    if cur_lo > target_range_lo + 24:
        print(
            f"[WARN] Footer start {cur_lo} is past expected band (~{target_range_lo}); cannot rewind.",
            flush=True,
        )
        return

    # One-shot jump when the paginator shows this page number (same index as script page #).
    if script_page > 0 and cur_lo < target_range_lo:
        if _try_click_numeric_page_button(driver, script_page):
            dismiss_applyboard_lightboxes(driver)
            time.sleep(0.25)
            r = _read_pagination_range(driver)
            if r:
                cur_lo = r[0]
                if cur_lo == target_range_lo or (target_range_lo <= cur_lo <= target_range_lo + 24):
                    print(
                        f"[Nav] Restored via page button {script_page} → {r[0]}-{r[1]} of {r[2]}.",
                        flush=True,
                    )
                    return

    print(
        f"[Nav] Restoring Students list to rows starting ~{target_range_lo} (was {cur_lo})…",
        flush=True,
    )
    # Budget: ~45s per Next step needed (14 steps from page 1 → page 15 can take several minutes).
    clicks_needed = max(1, (target_range_lo - cur_lo + 19) // 20)
    max_wall_s = min(960.0, max(200.0, float(clicks_needed) * 48.0))

    clicks = 0
    max_clicks = 80
    t_wall = time.monotonic()
    while cur_lo < target_range_lo and clicks < max_clicks:
        if time.monotonic() - t_wall > max_wall_s:
            print(
                f"[WARN] Pagination restore wall timeout ({max_wall_s:.0f}s) — retrying slower Next…",
                flush=True,
            )
            break
        if not go_next_page(driver, wait, err_dir, jump_fast=jump_fast):
            print("[WARN] Pagination restore stopped early.", flush=True)
            break
        clicks += 1
        r = _read_pagination_range(driver)
        if not r:
            break
        cur_lo = r[0]

    r = _read_pagination_range(driver)
    still_bad = bool(
        r
        and not (r[0] == target_range_lo or (target_range_lo <= r[0] <= target_range_lo + 24))
    )
    if still_bad and cur_lo < target_range_lo:
        print("[Nav] Restore finishing pass (standard Next, more reliable)…", flush=True)
        t_wall = time.monotonic()
        max_wall2 = 420.0
        while cur_lo < target_range_lo and clicks < max_clicks:
            if time.monotonic() - t_wall > max_wall2:
                break
            if not go_next_page(driver, wait, err_dir, jump_fast=False):
                break
            clicks += 1
            r = _read_pagination_range(driver)
            if not r:
                break
            cur_lo = r[0]
            if cur_lo == target_range_lo or (target_range_lo <= cur_lo <= target_range_lo + 24):
                break

    r = _read_pagination_range(driver)
    if r:
        ok = r[0] == target_range_lo or (target_range_lo <= r[0] <= target_range_lo + 24)
        if not ok:
            print(
                f"[WARN] Footer after restore: {r[0]}-{r[1]} of {r[2]} (wanted start ~{target_range_lo}).",
                flush=True,
            )


def _return_to_students_table(driver: webdriver.Chrome, err_dir: Path, list_url: str) -> None:
    """
    Return to the Students table after viewing a profile.
    Always uses driver.get(), never history.back() (Back pops to /agent in this SPA).
    Note: a bare /students load resets pagination to page 1 — callers must call
    ensure_students_list_footer_start afterward when continuing mid-list.
    """
    u = (list_url or "").strip()
    if u and "/students" in u and "applyboard.com" in u.lower():
        target = u.split("#")[0]
    else:
        target = _EFFECTIVE_STUDENTS_URL or STUDENTS_URL

    driver.get(target)
    wait_for_page(driver)
    _maybe_click_partners_splash(driver, err_dir=err_dir)
    _wait_students_list_ready(driver, err_dir=err_dir, timeout_s=180)


def _results_contains_student_id(results: List[StudentRow], student_id: str) -> bool:
    sid = (student_id or "").strip()
    if not sid:
        return False
    for r in results:
        if (r.student_id or "").strip() == sid:
            return True
    return False


def scrape_current_page(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    page_number: int,
    results: List[StudentRow],
    err_dir: Path,
    partial_save_every: int,
    out_dir: Path,
    max_students: Optional[int] = None,
    only_student_id: Optional[str] = None,
    skip_student_ids: Optional[Set[str]] = None,
    skip_positions_left: Optional[List[int]] = None,
    block_duplicate_ids: bool = False,
    pagination_footer_start: Optional[int] = None,
) -> int:
    # Return number of students processed on this page
    WebDriverWait(driver, 90).until(lambda d: len(get_student_data_rows(d)) > 0)
    processed = 0

    header_map = get_students_header_map(driver)
    if not header_map:
        print("[WARN] Could not read table headers; falling back to index-based parsing.", flush=True)

    rows = get_student_data_rows(driver)
    row_count = len(rows)
    rng = _read_pagination_range(driver)
    if rng:
        print(
            f"Found {row_count} rows | script page #{page_number} | "
            f"ApplyBoard range {rng[0]}-{rng[1]} of {rng[2]}.",
            flush=True,
        )
    else:
        print(f"Found {row_count} rows on script page #{page_number} (footer range not detected).", flush=True)

    for idx in range(row_count):
        if max_students is not None and processed >= max_students:
            break
        list_url = ""
        went_to_profile = False
        opened_new_tab = False
        try:
            # Re-fetch rows each loop to reduce stale element issues after navigation.
            rows = get_student_data_rows(driver)
            if idx >= len(rows):
                break

            row = rows[idx]
            cells = _row_cells(row)
            if len(cells) < 4:
                continue

            parsed = (
                parse_row_by_headers(cells, header_map) if header_map else parse_row_cells(cells)
            )
            student_id = (parsed.get("student_id") or "").strip()
            if not student_id:
                # Last resort: pull digits from the Student ID link text
                try:
                    for a in cells[1].find_elements(By.CSS_SELECTOR, "a"):
                        t = (a.text or "").strip()
                        if re.fullmatch(r"\d{4,}", t):
                            student_id = t
                            break
                except Exception:
                    pass
            if not student_id:
                m = re.search(r"\d{4,}", safe_text(cells[1]))
                if m:
                    student_id = m.group(0)
            if not student_id:
                for c in cells:
                    m = re.search(r"\b(\d{6,})\b", safe_text(c))
                    if m:
                        student_id = m.group(1)
                        break

            if not student_id:
                print(f"[SKIP] Row {idx+1} on page {page_number}: missing Student ID (likely header/filter row).")
                continue

            if only_student_id and student_id != only_student_id:
                continue

            if skip_positions_left is not None and skip_positions_left[0] > 0:
                skip_positions_left[0] -= 1
                print(
                    f"[RESUME] Skip row {idx + 1} page {page_number} (ID {student_id}) — "
                    f"{skip_positions_left[0]} position(s) left before scraping resumes.",
                    flush=True,
                )
                continue

            if skip_student_ids is not None and student_id in skip_student_ids:
                print(
                    f"[RESUME] Row {idx + 1} page {page_number}: skip Student ID {student_id} (in workbook, ID mode).",
                    flush=True,
                )
                continue

            if block_duplicate_ids and _results_contains_student_id(results, student_id):
                print(
                    f"[RESUME] Skip Student ID {student_id} — already in workbook (--resume-block-duplicate-ids).",
                    flush=True,
                )
                continue

            print(f"[Page {page_number}] ({idx+1}/{row_count}) Opening student {student_id}...", flush=True)

            list_url = driver.current_url or ""

            link_or_cell = get_student_link_from_row(row)
            if link_or_cell is None:
                raise NoSuchElementException("Could not find clickable element for student row")

            before_handles = list(driver.window_handles)
            opened_new_tab = _open_profile_in_new_tab(driver, link_or_cell, student_id=student_id)
            if opened_new_tab:
                # switch to newest tab
                new_handle = [h for h in driver.window_handles if h not in before_handles][-1]
                driver.switch_to.window(new_handle)
                went_to_profile = True
            else:
                before_url = driver.current_url or ""
                try:
                    click_with_retry(driver, link_or_cell)
                except Exception:
                    try:
                        driver.execute_script("arguments[0].click();", link_or_cell)
                    except Exception:
                        raise
                went_to_profile = True

            if not opened_new_tab:
                try:
                    _wait_after_student_click(driver, before_url=before_url)
                except TimeoutException:
                    print(f"[WARN] {student_id}: profile navigation wait timed out; extracting anyway...", flush=True)
                    time.sleep(1.2)
                except Exception:
                    time.sleep(0.8)

            phone, registration_date, profile_url, apps = scrape_student_profile(
                driver=driver,
                wait=wait,
                err_dir=err_dir,
                student_id=student_id,
            )

            results.append(
                StudentRow(
                    student_id=student_id or parsed["student_id"],
                    student_email=parsed["student_email"],
                    first_name=parsed["first_name"],
                    last_name=parsed["last_name"],
                    nationality=parsed["nationality"],
                    recruitment_partner=parsed["recruitment_partner"],
                    recruiter_type=parsed["recruiter_type"],
                    education=parsed["education"],
                    phone=phone,
                    registration_date=registration_date,
                    profile_url=profile_url,
                    page_number=page_number,
                    row_number=idx + 1,
                    destination_country=(apps.get("destination_country") or ""),
                    target_university=(apps.get("target_university") or ""),
                    target_program=(apps.get("target_program") or ""),
                    target_intake=(apps.get("target_intake") or ""),
                    city=(apps.get("city") or ""),
                    status_app=(apps.get("status_app") or ""),
                    intake_label=(apps.get("intake_label") or ""),
                )
            )

            # MySQL upsert (best-effort; keeps row unchanged if no change)
            try:
                if _MYSQL_SINK is not None:
                    _MYSQL_SINK.upsert_student(results[-1])
            except Exception as exc:
                print(f"[WARN] MySQL upsert failed for {student_id}: {exc}", flush=True)

            processed += 1
            applicant = f"{parsed.get('first_name', '')} {parsed.get('last_name', '')}".strip()
            print(
                f"[OK] Row {len(results)} | Page {page_number} | ID {student_id} | {applicant}",
                flush=True,
            )
            print(
                f"     Reg: {registration_date or '-'} | WhatsApp: {phone or '-'} | Email: {parsed.get('student_email', '') or '-'}",
                flush=True,
            )
            if apps:
                print(
                    f"     Dest: {apps.get('destination_country') or '-'} | App: {apps.get('status_app') or '-'} | Univ: {apps.get('target_university') or '-'} | Program: {apps.get('target_program') or '-'} | Intake: {apps.get('target_intake') or '-'}",
                    flush=True,
                )
            if partial_save_every > 0 and len(results) % partial_save_every == 0:
                save_results(results, out_dir, quiet=True)
                print(f"     >> Excel updated: {out_dir / OUTPUT_XLSX} ({len(results)} rows)", flush=True)

        except Exception as exc:
            print(f"[SKIP] Row {idx+1} on page {page_number}: {exc}", flush=True)
            screenshot_on_error(driver, err_dir, f"student_row_{page_number}_{idx+1}")
        finally:
            # Return to Students list
            if went_to_profile:
                try:
                    if opened_new_tab:
                        # Close profile tab and return without reloading Students
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        dismiss_applyboard_lightboxes(driver)
                        time.sleep(0.28)
                        # Avoid full reload unless grid is still empty (banner/overlays can hide rows briefly)
                        n_rows = len(get_student_data_rows(driver))
                        if n_rows == 0 and not _loose_students_grid_has_data(driver):
                            time.sleep(0.65)
                            dismiss_applyboard_lightboxes(driver)
                            n_rows = len(get_student_data_rows(driver))
                        if n_rows == 0 and not _loose_students_grid_has_data(driver):
                            _reload_students_list(driver, err_dir=err_dir)
                    else:
                        _return_to_students_table(driver, err_dir=err_dir, list_url=list_url)
                        dismiss_applyboard_lightboxes(driver)

                    if pagination_footer_start is not None:
                        ensure_students_list_footer_start(
                            driver,
                            wait,
                            err_dir,
                            pagination_footer_start,
                            jump_fast=True,
                            script_page=page_number,
                        )
                    time.sleep(0.12)
                except Exception:
                    pass

    return processed


def _pagination_is_first_or_last_button(btn: WebElement) -> bool:
    """Exclude << (first) and >> (last) — user wants only single-step next '>'."""
    try:
        label = (btn.get_attribute("aria-label") or "").lower()
        title = (btn.get_attribute("title") or "").lower()
        txt = (btn.text or "").strip()
        if "first" in label or "first" in title:
            return True
        if "last" in label or "last" in title:
            return True
        if txt in ("<<", ">>", "«", "»"):
            return True
    except Exception:
        pass
    return False


def _read_pagination_range(driver: webdriver.Chrome) -> Optional[Tuple[int, int, int]]:
    """
    Parse Students grid footer like '221 - 240 of 737' -> (221, 240, 737).

    The page body can contain several 'X - Y of Z' strings (other widgets). Prefer
    MUI table pagination nodes; otherwise take the match with the **largest total**
    so we never confuse the main grid with a smaller list.
    """
    pat = re.compile(r"\b(\d+)\s*[-–]\s*(\d+)\s+of\s+(\d+)\b", re.I)

    def _parse_best(text: str) -> Optional[Tuple[int, int, int]]:
        best: Optional[Tuple[int, int, int]] = None
        best_tot = -1
        for m in pat.finditer(text or ""):
            lo, hi, tot = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if tot > best_tot:
                best_tot = tot
                best = (lo, hi, tot)
        return best

    for sel in (
        ".MuiTablePagination-displayedRows",
        ".MuiTablePagination-toolbar",
        "[class*='TablePagination']",
        "nav[aria-label*='pagination' i]",
        "footer",
    ):
        try:
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                try:
                    if not el.is_displayed():
                        continue
                except Exception:
                    continue
                hit = _parse_best(el.text or "")
                if hit:
                    return hit
        except Exception:
            continue

    try:
        body = (driver.find_element(By.TAG_NAME, "body").text or "")
    except Exception:
        return None
    return _parse_best(body)


def _pagination_is_next_page_button(btn: WebElement) -> bool:
    """
    True for ONE step forward only: '>' / “Next page” / “Go to next page”.
    Never the double-arrow last-page control.
    """
    try:
        if not btn.is_displayed() or not btn.is_enabled():
            return False
        if _pagination_is_first_or_last_button(btn):
            return False
        label = (btn.get_attribute("aria-label") or "").lower()
        title = (btn.get_attribute("title") or "").lower()
        txt = (btn.text or "").strip()
        # Literal single chevron used on Students grid (37 pages)
        if txt in (">", "›", "→"):
            return True
        if "go to next page" in label or "go to next page" in title:
            return True
        if "next page" in label or "next page" in title:
            return True
        # MUI sometimes: aria-label="Go to next page"
        if label.strip() == "next page":
            return True
    except Exception:
        return False
    return False


def _pagination_is_next_page_button_loose(btn: WebElement) -> bool:
    """Same as _pagination_is_next_page_button but ignores disabled (for JS click retry)."""
    try:
        if not btn.is_displayed():
            return False
        if _pagination_is_first_or_last_button(btn):
            return False
        label = (btn.get_attribute("aria-label") or "").lower()
        title = (btn.get_attribute("title") or "").lower()
        txt = (btn.text or "").strip()
        if txt in (">", "›", "→"):
            return True
        if "go to next page" in label or "go to next page" in title:
            return True
        if "next page" in label or "next page" in title:
            return True
        if label.strip() == "next page":
            return True
    except Exception:
        return False
    return False


def _pagination_might_have_more(driver: webdriver.Chrome) -> bool:
    """True if footer suggests rows beyond the current window (e.g. 221-240 of 737)."""
    r = _read_pagination_range(driver)
    if r is None:
        return True
    return r[1] < r[2]


def go_next_page(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    err_dir: Path,
    *,
    jump_fast: bool = False,
) -> bool:
    """
    Clicks the single-step Next control (typically `>`), never `>>` (last page).
    Retries with JS click if needed; uses footer total so we do not stop early when
    other widgets embed smaller 'X-Y of Z' strings on the page.

    jump_fast: shorter waits, no verbose pagination dumps — use for resume bulk Next jumps.
    """
    if not _pagination_might_have_more(driver):
        if not jump_fast:
            print("[Nav] Footer shows end of list (no further pages).", flush=True)
        return False

    if jump_fast:
        if _go_next_page_single(driver, wait, err_dir, False, quiet=True, fast=True):
            return True
        if _go_next_page_single(driver, wait, err_dir, True, quiet=True, fast=True):
            return True
        if _pagination_might_have_more(driver):
            print(
                "[WARN] Fast Next failed on resume jump; retry without busy UI or increase waits.",
                flush=True,
            )
        return False

    for use_js in (False, True):
        for _attempt in range(2):
            ok = _go_next_page_single(driver, wait, err_dir, use_js, quiet=False, fast=False)
            if ok:
                return True
            time.sleep(0.55)

    if _pagination_might_have_more(driver):
        print(
            "[WARN] Next page control failed after retries, but the footer still shows more rows "
            "(e.g. not on last page). Check errors/ pagination_*.png or scroll the table footer into view.",
            flush=True,
        )
    return False


def _go_next_page_single(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    err_dir: Path,
    use_js: bool,
    *,
    quiet: bool = False,
    fast: bool = False,
) -> bool:
    """
    One attempt to go to the next page. use_js uses a programmatic click (MUI sometimes
    marks the chevron disabled while more pages exist).
    """
    wait_rows_s = 18 if fast else 45
    wait_changed_s = 28 if fast else 60
    ready_s = 55 if fast else 120
    settle_s = 0.1 if fast else 0.35
    wp_timeout = 12 if fast else DEFAULT_TIMEOUT_S

    try:
        WebDriverWait(driver, wait_rows_s).until(lambda d: len(get_student_data_rows(d)) > 0)
    except Exception:
        return False

    try:
        for sel in (".MuiTablePagination-root", "[class*='TablePagination']", "footer"):
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                try:
                    if el.is_displayed():
                        driver.execute_script(
                            "arguments[0].scrollIntoView({block:'nearest', inline:'nearest'});",
                            el,
                        )
                        break
                except Exception:
                    continue
            else:
                continue
            break
    except Exception:
        pass

    def _dump_pagination_debug(tag: str) -> None:
        if quiet:
            return
        try:
            screenshot_on_error(driver, err_dir, f"pagination_{tag}")
        except Exception:
            pass
        try:
            p = err_dir / f"{now_stamp()}_pagination_{tag}.html"
            p.write_text(driver.page_source or "", encoding="utf-8", errors="ignore")
        except Exception:
            pass
        try:
            # Print visible buttons summary (highest-signal)
            btns = driver.find_elements(By.CSS_SELECTOR, "button, a")
            rows: List[Tuple[int, str]] = []
            for b in btns:
                try:
                    if not b.is_displayed():
                        continue
                except Exception:
                    continue
                try:
                    tx = (b.text or "").strip().replace("\n", " ")
                except Exception:
                    tx = ""
                try:
                    al = (b.get_attribute("aria-label") or "").strip()
                except Exception:
                    al = ""
                try:
                    tt = (b.get_attribute("title") or "").strip()
                except Exception:
                    tt = ""
                try:
                    dis = (b.get_attribute("disabled") or "").strip()
                except Exception:
                    dis = ""
                try:
                    en = b.is_enabled()
                except Exception:
                    en = False
                score = 0
                low = (al + " " + tt).lower()
                if "next page" in low or "go to next page" in low:
                    score += 20
                if tx in (">", "›", "→"):
                    score += 15
                if tx in (">>", "»", "<<", "«"):
                    score -= 10
                if not en or dis:
                    score -= 3
                if tx:
                    score += 1
                rows.append((score, f"btn txt='{tx}' aria='{al}' title='{tt}' enabled={en} disabled_attr='{dis}'"))
            rows.sort(key=lambda x: -x[0])
            print("[DBG] Pagination visible buttons (top 25):", flush=True)
            for _, line in rows[:25]:
                print("      " + line, flush=True)
        except Exception:
            pass

    candidates: List[WebElement] = []

    # 1) Most reliable (matches MUI paginator): aria-label "Go to next page"
    try:
        candidates.extend(driver.find_elements(By.CSS_SELECTOR, "button[aria-label*='go to next page' i]"))
        candidates.extend(driver.find_elements(By.CSS_SELECTOR, "a[aria-label*='go to next page' i]"))
    except Exception:
        pass

    # 2) Generic next page aria-labels
    try:
        candidates.extend(driver.find_elements(By.CSS_SELECTOR, "button[aria-label*='next page' i]"))
        candidates.extend(driver.find_elements(By.CSS_SELECTOR, "a[aria-label*='next page' i]"))
    except Exception:
        pass

    # 3) If the control actually has a literal ">" text (sometimes does)
    try:
        candidates.extend(driver.find_elements(By.XPATH, "//button[normalize-space()='>']"))
        candidates.extend(driver.find_elements(By.XPATH, "//a[normalize-space()='>']"))
    except Exception:
        pass

    # 4) Fallback: scan likely pagination containers and collect clickable elements
    try:
        for root in driver.find_elements(
            By.CSS_SELECTOR,
            "nav, .MuiTablePagination-root, .MuiPagination-root, [class*='TablePagination'], [class*='pagination'], footer, [role='navigation']",
        ):
            try:
                candidates.extend(root.find_elements(By.CSS_SELECTOR, "button, a"))
            except Exception:
                pass
    except Exception:
        pass

    # 5) Strong fallback for your exact UI: click the button immediately BEFORE the "last page" control (>>).
    try:
        last_btns = driver.find_elements(
            By.CSS_SELECTOR,
            "button[aria-label*='last' i], button[title*='last' i], a[aria-label*='last' i], a[title*='last' i]",
        )
        for lb in last_btns:
            try:
                if not lb.is_displayed():
                    continue
            except Exception:
                continue
            try:
                parent = lb.find_element(By.XPATH, "..")
                sibs = parent.find_elements(By.CSS_SELECTOR, "button, a")
                if len(sibs) >= 2:
                    # Pick the immediate previous sibling in the same parent list
                    idx = sibs.index(lb)
                    if idx > 0:
                        candidates.append(sibs[idx - 1])
            except Exception:
                continue
    except Exception:
        pass

    seen: set[int] = set()
    uniq_candidates: List[WebElement] = []
    for b in candidates:
        try:
            k = id(b)
            if k in seen:
                continue
            seen.add(k)
            uniq_candidates.append(b)
        except Exception:
            continue

    filtered = [b for b in uniq_candidates if _pagination_is_next_page_button(b)]
    if not filtered and use_js:
        filtered = [b for b in uniq_candidates if _pagination_is_next_page_button_loose(b)]

    def _click_next_via_range_label() -> Optional[WebElement]:
        """
        Last-resort for the exact Students paginator you showed:
        find the '1 - 20 of 737' label, then click the single-step next button which is
        typically the button immediately BEFORE the last-page button in that same paginator.
        """
        try:
            # Find candidate label nodes
            labels = driver.find_elements(
                By.XPATH,
                "//*[contains(normalize-space(.), 'of') and contains(normalize-space(.), '-') and string-length(normalize-space(.)) < 40]",
            )
        except Exception:
            labels = []

        best = None
        for el in labels:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                continue
            try:
                t = (el.text or "").strip()
            except Exception:
                continue
            if not re.search(r"\b\d+\s*-\s*\d+\s+of\s+\d+\b", t, re.I):
                continue
            best = el
            break

        if best is None:
            return None

        # Walk up to find a container with multiple pagination buttons.
        node = best
        for _ in range(8):
            try:
                node = node.find_element(By.XPATH, "..")
            except Exception:
                break
            try:
                btns = node.find_elements(By.CSS_SELECTOR, "button, a")
            except Exception:
                continue
            visible = []
            for b in btns:
                try:
                    if b.is_displayed():
                        visible.append(b)
                except Exception:
                    continue
            if len(visible) >= 6:
                # Guess: last-page is the last visible control, next is previous
                # Choose the last enabled element that is not the last-page control.
                for cand in reversed(visible):
                    try:
                        if _pagination_is_first_or_last_button(cand):
                            continue
                        if not cand.is_enabled():
                            continue
                        # Prefer literal single chevron
                        tx = (cand.text or "").strip()
                        if tx in (">", "›", "→"):
                            return cand
                    except Exception:
                        continue
                # Otherwise take the element immediately before any 'last' control
                for i, cand in enumerate(visible):
                    try:
                        if _pagination_is_first_or_last_button(cand):
                            if i > 0 and visible[i - 1].is_enabled():
                                return visible[i - 1]
                    except Exception:
                        continue
                # Fallback: second last enabled
                enabled = [b for b in visible if getattr(b, "is_enabled", lambda: False)()]
                if len(enabled) >= 2:
                    return enabled[-2]
        return None

    if not filtered:
        cand = _click_next_via_range_label()
        if cand is not None:
            filtered = [cand]

    if not filtered and use_js:
        for sel in (
            "button[aria-label*='go to next page' i]",
            "button[aria-label*='next page' i]",
            "a[aria-label*='go to next page' i]",
        ):
            try:
                for b in driver.find_elements(By.CSS_SELECTOR, sel):
                    if _pagination_is_next_page_button_loose(b):
                        filtered = [b]
                        break
                if filtered:
                    break
            except Exception:
                continue

    if not filtered:
        _dump_pagination_debug("no_next_found")
        return False

    def _next_btn_rank(elem: WebElement) -> int:
        try:
            tx = (elem.text or "").strip()
            if tx == ">":
                return 4
            al = (elem.get_attribute("aria-label") or "").lower()
            if "next page" in al:
                return 3
            return 2
        except Exception:
            return 0

    filtered.sort(key=_next_btn_rank, reverse=True)
    btn = filtered[0]

    old_url = driver.current_url
    old_rows = get_student_data_rows(driver)
    old_first_row_text = safe_text(old_rows[0]) if old_rows else ""
    old_rng = _read_pagination_range(driver)

    try:
        if use_js:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.25)
            driver.execute_script("arguments[0].click();", btn)
        else:
            click_with_retry(driver, btn)
    except Exception:
        _dump_pagination_debug("click_failed")
        return False

    # Wait for either URL change OR first data row content change.
    def changed(d: webdriver.Chrome) -> bool:
        try:
            new_url = d.current_url or ""
            if new_url != (old_url or ""):
                return True
            new_rows = get_student_data_rows(d)
            if not new_rows:
                return False
            return safe_text(new_rows[0]) != old_first_row_text
        except Exception:
            return False

    try:
        WebDriverWait(driver, wait_changed_s).until(changed)
    except TimeoutException:
        # Could still be last page or slow React; treat as not navigated.
        return False

    wait_for_page(driver, timeout_s=wp_timeout)
    try:
        _wait_students_list_ready(driver, err_dir=err_dir, timeout_s=ready_s)
    except Exception:
        pass
    time.sleep(settle_s)

    new_rng = _read_pagination_range(driver)
    new_rows = get_student_data_rows(driver)
    new_first = safe_text(new_rows[0]) if new_rows else ""

    # Confirm the grid window advanced (guards false-positive "change" when pagination stuck).
    if new_first != old_first_row_text:
        # First row updated — trust navigation even if footer regex matched the wrong widget.
        if new_rng:
            nav_msg = (
                f"[Nav] {new_rng[0]}-{new_rng[1]} / {new_rng[2]}"
                if fast
                else f"[Nav] Table range: {new_rng[0]}-{new_rng[1]} of {new_rng[2]}."
            )
            print(nav_msg, flush=True)
        return True

    if old_rng and new_rng and old_rng[2] == new_rng[2]:
        if new_rng[0] <= old_rng[0] and new_first == old_first_row_text:
            if not quiet:
                print(
                    f"[WARN] Next page did not move range (still {old_rng[0]}-{old_rng[1]} of {old_rng[2]}).",
                    flush=True,
                )
            _dump_pagination_debug("range_unchanged")
            return False

    if new_rng:
        nav_msg = (
            f"[Nav] {new_rng[0]}-{new_rng[1]} / {new_rng[2]}" if fast else f"[Nav] Table range: {new_rng[0]}-{new_rng[1]} of {new_rng[2]}."
        )
        print(nav_msg, flush=True)
    return True


def _xlsx_cell_str(val: Any) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):  # type: ignore[arg-type]
            return ""
    except Exception:
        pass
    return str(val).strip()


def load_student_rows_from_xlsx(xlsx_path: Path) -> List[StudentRow]:
    """
    Reconstruct StudentRow list from a prior applyboard_students.xlsx (for --resume).
    Rows without Student ID are skipped.
    """
    if not xlsx_path.exists():
        return []
    try:
        df = pd.read_excel(xlsx_path)
    except Exception:
        return []
    if df is None or df.empty:
        return []
    out: List[StudentRow] = []
    for _, ser in df.iterrows():
        sid = _xlsx_cell_str(ser.get("Student ID"))
        if not sid:
            continue
        name = _xlsx_cell_str(ser.get("Applicant Name"))
        parts = name.split(None, 1) if name else []
        first = parts[0] if parts else ""
        last = parts[1] if len(parts) > 1 else ""
        out.append(
            StudentRow(
                student_id=sid,
                student_email=_xlsx_cell_str(ser.get("Applicant Email")),
                first_name=first,
                last_name=last,
                nationality=_xlsx_cell_str(ser.get("Country")),
                education=_xlsx_cell_str(ser.get("Education Level")),
                phone=_xlsx_cell_str(ser.get("WhatsApp Num")),
                registration_date=_xlsx_cell_str(ser.get("Registration Date")),
                destination_country=_xlsx_cell_str(ser.get("Destination/Country")),
                target_university=_xlsx_cell_str(ser.get("Target University")),
                target_program=_xlsx_cell_str(ser.get("Target Program")),
                target_intake=_xlsx_cell_str(ser.get("Target Intake")),
                city=_xlsx_cell_str(ser.get("City")),
                status_app=_xlsx_cell_str(ser.get("Status App")),
                intake_label=_xlsx_cell_str(ser.get("In-take")),
                graduation_notes=_xlsx_cell_str(ser.get("Graduation / Notes")),
            )
        )
    return out


def dedupe_student_rows_keep_first_inplace(results: List[StudentRow]) -> int:
    """
    Collapse duplicate Student IDs (keep first row). Returns number of rows removed.
    """
    seen: Set[str] = set()
    keep: List[StudentRow] = []
    dupes = 0
    for r in results:
        sid = (r.student_id or "").strip()
        if not sid:
            keep.append(r)
            continue
        if sid in seen:
            dupes += 1
            continue
        seen.add(sid)
        keep.append(r)
    results[:] = keep
    return dupes


def save_results(results: List[StudentRow], out_dir: Path, quiet: bool = False) -> None:
    """Writes Excel + CSV using EXCEL_COLUMNS order (your template)."""
    nd = dedupe_student_rows_keep_first_inplace(results)
    if nd > 0 and not quiet:
        print(
            f"[WARN] Removed {nd} duplicate row(s) by Student ID before save (kept first occurrence).",
            flush=True,
        )
    if not results:
        df = pd.DataFrame(columns=EXCEL_COLUMNS)
    else:
        df = pd.DataFrame([r.to_excel_row(i) for i, r in enumerate(results, start=1)])
        df = df.reindex(columns=EXCEL_COLUMNS)

    xlsx_path = out_dir / OUTPUT_XLSX
    csv_path = out_dir / OUTPUT_CSV

    try:
        df.to_excel(xlsx_path, index=False)
    except PermissionError:
        # Common on Windows when the file is open in Excel.
        alt = out_dir / f"applyboard_students_{now_stamp()}.xlsx"
        df.to_excel(alt, index=False)
        xlsx_path = alt
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # Excel formatting (openpyxl)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Header style
        header_font = Font(bold=True, size=13, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E6DEB")  # ApplyBoard-ish blue
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_col = ws.max_column
        max_row = ws.max_row

        ws.freeze_panes = "A2"
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        # Wrap long-text columns
        wrap_cols = {
            "Applicant Name",
            "Target University",
            "Target Program",
            "Target Intake",
            "Education Level",
            "Graduation / Notes",
        }
        wrap_col_idxs = {EXCEL_COLUMNS.index(c) + 1 for c in wrap_cols if c in EXCEL_COLUMNS}

        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.row_dimensions[1].height = 24

        # Data alignment
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                wrap = c in wrap_col_idxs
                ws.cell(row=r, column=c).alignment = Alignment(
                    vertical="top",
                    wrap_text=wrap,
                )

        # Column widths (auto-fit with cap; long content wraps)
        MIN_W = 10
        MAX_W = 45
        for c in range(1, max_col + 1):
            header = (ws.cell(row=1, column=c).value or "")
            if header in ("Target Program", "Graduation / Notes"):
                ws.column_dimensions[get_column_letter(c)].width = 45
                continue
            if header in ("Applicant Email",):
                ws.column_dimensions[get_column_letter(c)].width = 28
                continue

            max_len = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v)
                # Only consider first line for width; wrapping handles the rest
                s = s.splitlines()[0]
                max_len = max(max_len, len(s))

            width = min(MAX_W, max(MIN_W, max_len + 2))
            ws.column_dimensions[get_column_letter(c)].width = width

        wb.save(xlsx_path)
    except Exception:
        # Formatting is best-effort; keep CSV/Excel data even if styling fails.
        pass

    if not quiet:
        print(f"Saved {len(results)} row(s) -> {xlsx_path} | {csv_path}", flush=True)


def build_driver(*, headless: bool = False) -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    # Fewer renderer/chrome crashes during long Selenium runs (safe on Windows/Linux).
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-features=TranslateUI")
    # Linux VPS + running as root: Chrome will exit unless sandbox is disabled.
    if sys.platform.startswith("linux"):
        try:
            if hasattr(os, "geteuid") and os.geteuid() == 0:  # type: ignore[attr-defined]
                options.add_argument("--no-sandbox")
                options.add_argument("--disable-setuid-sandbox")
        except Exception:
            pass
        # Reduce GPU/renderer assumptions on servers.
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--disable-extensions")
        options.add_argument("--no-first-run")
        options.add_argument("--no-default-browser-check")

    # Allow overriding Chrome binary location (useful when only chromium-browser exists).
    chrome_bin = (os.getenv("CHROME_BINARY") or "").strip()
    if chrome_bin:
        options.binary_location = chrome_bin
    if headless:
        # Chrome's newer headless; works without X11/Wayland (typical Linux VPS).
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
    else:
        options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    try:
        driver.set_page_load_timeout(120)
    except Exception:
        pass
    return driver


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    out_dir, err_dir = ensure_dirs(base_dir)

    # Load .env if present (so credentials can be provided without prompts)
    load_dotenv(dotenv_path=base_dir / ".env", override=False)

    parser = argparse.ArgumentParser(description="Scrape ApplyBoard Students to Excel/CSV.")
    parser.add_argument("--email", help="ApplyBoard email (overrides env/prompt).")
    parser.add_argument("--password", help="ApplyBoard password (overrides env/prompt). WARNING: visible in shell history.")
    parser.add_argument("--max-pages", type=int, default=0, help="Max pages to scrape (0 = all).")
    parser.add_argument("--max-students", type=int, default=0, help="Max students per page to scrape (0 = all). Useful for debugging.")
    parser.add_argument("--only-student-id", default="", help="Scrape only this Student ID (debug).")
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Load existing Excel for merge, then continue: skip the first N rows in list order (N = row count in file unless --resume-after).",
    )
    parser.add_argument(
        "--resume-from",
        default="",
        metavar="PATH",
        help=f"Workbook to load for --resume (default: {OUTPUT_XLSX} next to this script).",
    )
    parser.add_argument(
        "--resume-after",
        type=int,
        default=None,
        metavar="N",
        help="With --resume: skip the first N rows of the table (no profile open). Default: number of rows loaded from Excel.",
    )
    parser.add_argument(
        "--resume-skip-saved-ids",
        action="store_true",
        help="Also skip rows whose Student ID is in the workbook (legacy; can mis-skip if pagination is wrong).",
    )
    parser.add_argument(
        "--resume-block-duplicate-ids",
        action="store_true",
        help="Skip profiles when Student ID is already in the workbook (off by default — sort changes cause false skips).",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=20,
        metavar="N",
        help="Rows per Students table page (ApplyBoard default 20). Used for smart resume jump.",
    )
    parser.add_argument(
        "--resume-start-page",
        type=int,
        default=0,
        metavar="N",
        help="With --resume: jump using Next-only to page N (overrides row-count÷page-size). Match --resume-after.",
    )
    parser.add_argument(
        "--start-page",
        type=int,
        default=1,
        metavar="N",
        help="Without smart resume: click Next until page N. Smart resume uses --resume-start-page or row math.",
    )
    parser.add_argument(
        "--mysql-db-php",
        default="",
        help=(
            "Path to PHP db config (db.php). If omitted: uses APPLYBOARD_MYSQL_DB_PHP, or ./db.php next to the script, "
            "or Windows default XAMPP db.php."
        ),
    )
    parser.add_argument("--mysql-host", default="", help="MySQL host (overrides db.php). Env: APPLYBOARD_MYSQL_HOST")
    parser.add_argument("--mysql-db", default="", help="MySQL database name (overrides db.php). Env: APPLYBOARD_MYSQL_DB")
    parser.add_argument("--mysql-user", default="", help="MySQL username (overrides db.php). Env: APPLYBOARD_MYSQL_USER")
    parser.add_argument(
        "--mysql-password",
        default="",
        help="MySQL password (overrides db.php). Env: APPLYBOARD_MYSQL_PASSWORD",
    )
    parser.add_argument("--mysql-port", type=int, default=0, help="MySQL port (overrides db.php). Env: APPLYBOARD_MYSQL_PORT")
    parser.add_argument(
        "--mysql-table",
        default="applyboard_students",
        help="MySQL table name for upserts.",
    )
    parser.add_argument(
        "--no-mysql",
        action="store_true",
        help="Disable MySQL upserts.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Chrome headless. On Windows this is the only way to enable headless (APPLYBOARD_HEADLESS in .env is ignored).",
    )
    parser.add_argument(
        "--headful",
        action="store_true",
        help="Force visible Chrome on Linux/macOS (overrides auto headless when DISPLAY is unset).",
    )
    parser.add_argument(
        "--login-url",
        default="",
        metavar="URL",
        help=(
            "Login entry URL. Default: partners /login. "
            "Use https://id.applyboard.com/ for ApplyBoard Identity (same modal as your screenshot). "
            "Prefer APPLYBOARD_LOGIN_URL in .env; do not paste links with state=/nonce= (one-time OAuth)."
        ),
    )
    args = parser.parse_args()

    resume_xlsx = (
        Path(args.resume_from).expanduser()
        if (args.resume_from or "").strip()
        else (base_dir / OUTPUT_XLSX)
    )

    email = (args.email or os.getenv("APPLYBOARD_EMAIL") or "").strip()
    password = args.password or os.getenv("APPLYBOARD_PASSWORD") or ""
    # Windows: only --headless (keeps .env from hiding the browser during local testing).
    # Linux/macOS: APPLYBOARD_HEADLESS=1, or Linux with no DISPLAY (typical VPS) → headless; use --headful or APPLYBOARD_HEADFUL=1 to show UI.
    headless = bool(args.headless)
    if sys.platform != "win32":
        headless = headless or (
            os.getenv("APPLYBOARD_HEADLESS", "").strip().lower() in ("1", "true", "yes", "on")
        )
        if sys.platform.startswith("linux") and not (os.getenv("DISPLAY") or "").strip():
            headless = True
        if os.getenv("APPLYBOARD_HEADFUL", "").strip().lower() in ("1", "true", "yes", "on"):
            headless = False
        if args.headful:
            headless = False

    if headless:
        print("Chrome running headless (no UI).", flush=True)

    if not email:
        email = input("ApplyBoard email: ").strip()
    if not password:
        password = getpass("ApplyBoard password: ")

    results: List[StudentRow] = []
    driver: Optional[webdriver.Chrome] = None

    try:
        driver = build_driver(headless=headless)
        wait = WebDriverWait(driver, DEFAULT_TIMEOUT_S)

        # Initialize MySQL sink (best-effort)
        global _MYSQL_SINK
        if not args.no_mysql:
            host = (args.mysql_host or os.getenv("APPLYBOARD_MYSQL_HOST") or "").strip()
            dbn = (args.mysql_db or os.getenv("APPLYBOARD_MYSQL_DB") or "").strip()
            usr = (args.mysql_user or os.getenv("APPLYBOARD_MYSQL_USER") or "").strip()
            pw = args.mysql_password or os.getenv("APPLYBOARD_MYSQL_PASSWORD") or ""
            port = int(args.mysql_port or int(os.getenv("APPLYBOARD_MYSQL_PORT", "0") or "0") or 0)

            cfg: Optional[Dict[str, str]] = None
            if host and dbn and usr:
                cfg = {"host": host, "db": dbn, "user": usr, "password": pw, "port": str(port or 3306)}
            else:
                php_path = effective_mysql_db_php_path(base_dir=base_dir, cli_value=args.mysql_db_php)
                cfg = _parse_php_db_config(php_path)
                if cfg is None:
                    print(
                        f"[WARN] MySQL disabled (no config). Provide --mysql-host/--mysql-db/--mysql-user "
                        f"or a db.php via --mysql-db-php / APPLYBOARD_MYSQL_DB_PHP. Tried: {php_path}",
                        flush=True,
                    )
            if cfg:
                try:
                    _MYSQL_SINK = MySQLSink(
                        host=cfg["host"],
                        db=cfg["db"],
                        user=cfg["user"],
                        password=cfg.get("password", ""),
                        port=int(cfg.get("port", "3306") or "3306"),
                    )
                    _MYSQL_SINK.ensure_table(table=args.mysql_table)
                    print(f"MySQL upsert enabled -> {cfg['db']}.{args.mysql_table}", flush=True)
                except Exception as exc:
                    _MYSQL_SINK = None
                    print(f"[WARN] MySQL disabled (init failed): {exc}", flush=True)

        login(driver, wait, email, password, err_dir, login_entry_url=effective_login_url(args.login_url))

        print("Opening Students page...")
        wait_for_students_table(driver, wait, err_dir)

        skip_positions_left: List[int] = [0]
        skip_ids: Set[str] = set()
        page_size = max(5, min(100, int(args.page_size or 20)))

        if args.resume:
            if resume_xlsx.exists():
                results = load_student_rows_from_xlsx(resume_xlsx)
                n_loaded = len(results)
                n_skip = args.resume_after if args.resume_after is not None else n_loaded
                skip_positions_left[0] = max(0, int(n_skip))
                if args.resume_skip_saved_ids:
                    skip_ids = {r.student_id for r in results if (r.student_id or "").strip()}
                print(
                    f"Resume: loaded {n_loaded} row(s) from {resume_xlsx.name}; "
                    f"resume row offset N={skip_positions_left[0]}.",
                    flush=True,
                )
                if args.resume_skip_saved_ids:
                    print(
                        f"         Also using ID skip for {len(skip_ids)} Student ID(s) (--resume-skip-saved-ids).",
                        flush=True,
                    )
            else:
                print(f"[WARN] --resume but file not found: {resume_xlsx}. Starting fresh.", flush=True)
                results = []
                save_results([], out_dir)
                print(f"Excel file created with headers: {out_dir / OUTPUT_XLSX}", flush=True)
        else:
            results = []
            if args.resume_after is not None:
                skip_positions_left[0] = max(0, int(args.resume_after))
            save_results([], out_dir)
            print(f"Excel file created with headers: {out_dir / OUTPUT_XLSX}", flush=True)

        resume_smart_jump = False
        if args.resume and resume_xlsx.exists() and skip_positions_left[0] > 0:
            n_skip = skip_positions_left[0]
            if int(args.resume_start_page or 0) > 0:
                full_pages = max(0, int(args.resume_start_page) - 1)
                print(
                    f"Resume: using --resume-start-page {args.resume_start_page} → {full_pages} Next-only jump(s).",
                    flush=True,
                )
            else:
                full_pages = n_skip // page_size
            remainder = n_skip % page_size
            skip_positions_left[0] = remainder

            print(
                f"Resume: smart jump — {full_pages}× Next (no row opens), "
                f"then skip {remainder} row(s) on this page (page size {page_size}). "
                f"Target script page {full_pages + 1} after jump.",
                flush=True,
            )

            jumped = 0
            for _ in range(full_pages):
                if not go_next_page(driver, wait, err_dir, jump_fast=True):
                    print(
                        f"[WARN] Resume jump stopped after {jumped}/{full_pages} successful Next click(s). "
                        f"Continuing from current page; adjust --resume-start-page / --resume-after if needed.",
                        flush=True,
                    )
                    break
                jumped += 1

            page = jumped + 1
            resume_smart_jump = True

            rng_j = _read_pagination_range(driver)
            if rng_j:
                exp_lo = full_pages * page_size + 1
                print(
                    f"Resume: footer now {rng_j[0]}-{rng_j[1]} of {rng_j[2]} "
                    f"(expect range start ~{exp_lo} if sort unchanged).",
                    flush=True,
                )

        if not resume_smart_jump:
            effective_start_page = max(1, int(args.start_page or 1))
            if effective_start_page > 1:
                print(f"Advancing to start page {effective_start_page}...", flush=True)
                ok = True
                for p in range(1, effective_start_page):
                    if not go_next_page(driver, wait, err_dir):
                        print(
                            f"[WARN] Could not advance to page {effective_start_page} (failed after page {p}). "
                            f"Continue from current page or fix filters/pagination.",
                            flush=True,
                        )
                        ok = False
                        break
                if ok:
                    print(f"Start page {effective_start_page} reached.", flush=True)
            page = effective_start_page
        while True:
            print(f"\n=== Scraping page {page} ===")
            scraped = scrape_current_page(
                driver=driver,
                wait=wait,
                page_number=page,
                results=results,
                err_dir=err_dir,
                partial_save_every=1,
                out_dir=out_dir,
                max_students=(args.max_students or None),
                only_student_id=(args.only_student_id.strip() or None),
                skip_student_ids=skip_ids if skip_ids else None,
                skip_positions_left=skip_positions_left if skip_positions_left[0] > 0 else None,
                block_duplicate_ids=bool(args.resume_block_duplicate_ids),
                pagination_footer_start=(page - 1) * page_size + 1,
            )

            if scraped == 0:
                print("No rows processed on this page (may be empty or selector changed).")

            save_results(results, out_dir)

            print("Trying next page...")
            if args.max_pages and page >= args.max_pages:
                print(f"Reached --max-pages={args.max_pages}. Stopping pagination.")
                break

            if not go_next_page(driver, wait, err_dir):
                print("No next page detected. Finished pagination.")
                break

            page += 1

        print("\nDone.")
        save_results(results, out_dir)

    except InvalidSessionIdException as exc:
        print(
            "\n[FATAL] Browser session ended (Chrome was closed, crashed, or the driver disconnected).\n"
            "        Keep the Chrome window that automation opens visible until the script finishes;\n"
            "        do not close it manually. Update Chrome and chromedriver if crashes repeat.\n",
            flush=True,
        )
        if results:
            save_results(results, out_dir)
        raise exc
    except Exception as exc:
        print(f"Fatal error: {exc}")
        if driver is not None:
            screenshot_on_error(driver, err_dir, "fatal")
        if results:
            save_results(results, out_dir)
        raise
    finally:
        try:
            if _MYSQL_SINK is not None:
                _MYSQL_SINK.close()
        except Exception:
            pass
        _safe_quit_driver(driver)


if __name__ == "__main__":
    main()