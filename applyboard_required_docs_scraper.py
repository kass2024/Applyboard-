from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from getpass import getpass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Reuse the battle-tested login + driver config from the main scraper.
import applyboard_scraper as ab


OUTPUT_CSV = "applyboard_required_documents.csv"
OUTPUT_XLSX = "applyboard_required_documents.xlsx"


@dataclass
class RequirementRow:
    student_id: str
    app_index: int
    school: str
    program: str
    intake: str
    status: str
    requirements: List[Dict[str, str]]


def _safe_text(el: Optional[WebElement]) -> str:
    if el is None:
        return ""
    try:
        return (el.text or "").strip()
    except Exception:
        return ""


def _first_visible(driver: ab.webdriver.Chrome, xpaths: List[str]) -> Optional[WebElement]:
    for xp in xpaths:
        try:
            el = driver.find_element(By.XPATH, xp)
            if el and el.is_displayed():
                return el
        except Exception:
            continue
    return None


def _open_in_new_tab(driver: ab.webdriver.Chrome, el: WebElement) -> bool:
    """
    Prefer opening via href in a new tab; fall back to click and wait for handle count change.
    """
    before = list(driver.window_handles)
    href = ""
    try:
        href = (el.get_attribute("href") or "").strip()
    except Exception:
        href = ""
    if href:
        try:
            driver.execute_script("window.open(arguments[0], '_blank');", href)
            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(before))
            return True
        except Exception:
            pass
    try:
        ab.click_with_retry(driver, el, timeout_s=12, retries=3)
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(before))
        return True
    except Exception:
        return False


def _extract_view_url(view_el: WebElement) -> str:
    """
    Paid Applications 'View' is sometimes an <a>, sometimes a <button> that wraps/contains an <a>.
    Try best-effort to find a stable href.
    """
    try:
        href = (view_el.get_attribute("href") or "").strip()
        if href:
            return href
    except Exception:
        pass
    # look for nearest ancestor <a>
    try:
        a = view_el.find_element(By.XPATH, "./ancestor::a[1]")
        href = (a.get_attribute("href") or "").strip()
        if href:
            return href
    except Exception:
        pass
    # look for descendant <a>
    try:
        a = view_el.find_element(By.XPATH, ".//a[1]")
        href = (a.get_attribute("href") or "").strip()
        if href:
            return href
    except Exception:
        pass
    return ""


def _switch_to_newest_tab(driver: ab.webdriver.Chrome, before_handles: List[str]) -> Optional[str]:
    after = list(driver.window_handles)
    for h in after:
        if h not in before_handles:
            try:
                driver.switch_to.window(h)
                return h
            except Exception:
                continue
    # fallback: last handle
    try:
        driver.switch_to.window(after[-1])
        return after[-1]
    except Exception:
        return None


def _close_tab_and_return(driver: ab.webdriver.Chrome, handle: str, return_handle: str) -> None:
    try:
        driver.switch_to.window(handle)
        driver.close()
    except Exception:
        pass
    try:
        driver.switch_to.window(return_handle)
    except Exception:
        pass


def _find_paid_app_view_buttons(driver: ab.webdriver.Chrome) -> List[Tuple[WebElement, WebElement]]:
    """
    Return list of (view_button, row_container) in the "Paid Applications" section.
    """
    out: List[Tuple[WebElement, WebElement]] = []
    try:
        secs = driver.find_elements(By.XPATH, "//section[.//*[contains(normalize-space(.), 'Paid Applications')]]")
    except Exception:
        secs = []
    if not secs:
        # Some accounts render without <section>
        try:
            secs = driver.find_elements(
                By.XPATH,
                "//*[self::div or self::main][.//*[contains(normalize-space(.), 'Paid Applications')]]",
            )
        except Exception:
            secs = []

    for sec in secs[:3]:
        try:
            if not sec.is_displayed():
                continue
        except Exception:
            pass
        try:
            views = sec.find_elements(By.XPATH, ".//*[self::a or self::button][contains(normalize-space(.), 'View')]")
        except Exception:
            views = []
        for vb in views:
            try:
                if not vb.is_displayed():
                    continue
            except Exception:
                continue
            row = None
            # best row containers
            for xp in [
                "./ancestor::*[@role='row'][1]",
                "./ancestor::tr[1]",
                "./ancestor::*[self::div or self::li][1]",
            ]:
                try:
                    row = vb.find_element(By.XPATH, xp)
                    if row is not None:
                        break
                except Exception:
                    continue
            if row is None:
                continue
            out.append((vb, row))
    return out


def _extract_paid_row_metadata(row_el: WebElement) -> Dict[str, str]:
    """
    Best-effort extract of School / Program / Intake / Status from the Paid Applications row text.
    """
    txt = (_safe_text(row_el) or "").replace("\n", "\n").strip()
    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
    low_lines = [ln.lower() for ln in lines]

    # School: sometimes in an <img alt="...">
    school = ""
    try:
        for img in row_el.find_elements(By.CSS_SELECTOR, "img[alt]"):
            alt = (img.get_attribute("alt") or "").strip()
            if alt and len(alt) >= 3 and not re.search(r"\b(applyboard|logo)\b", alt, re.I):
                school = alt
                break
    except Exception:
        pass

    # Fallback: first non-trivial line that looks like a school name (not status/app#/view)
    noise = {"view", "requirements", "paid applications", "priority"}
    if not school:
        for ln, lln in zip(lines, low_lines):
            if lln in noise:
                continue
            if re.fullmatch(r"\d{1,4}", ln):
                continue
            if "processing" in lln or "submitted" in lln or "accepted" in lln or "rejected" in lln:
                continue
            if len(ln) >= 4:
                school = ln
                break

    # Program: find a "Bachelor of / Master of" style or long line containing hyphen + field
    program = ""
    for ln, lln in zip(lines, low_lines):
        if lln in noise or lln == school.lower():
            continue
        if "bachelor" in lln or "master" in lln or "diploma" in lln or "certificate" in lln:
            program = ln
            break
    if not program:
        for ln, lln in zip(lines, low_lines):
            if lln in noise or lln == school.lower():
                continue
            if len(ln) >= 10 and not re.fullmatch(r"\d{1,4}", ln):
                program = ln
                break

    # Intake: typical "2026-Sep" or "Sep 2026"
    intake = ""
    m = re.search(r"\b(20\d{2})-(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b", txt, re.I)
    if m:
        intake = f"{m.group(1)}-{m.group(2).title()}"
    if not intake:
        m = re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+(20\d{2})\b", txt, re.I)
        if m:
            intake = f"{m.group(2)}-{m.group(1).title()}"

    # Status: reuse helper from main scraper
    status = ""
    try:
        status = ab._primary_paid_application_status(txt)  # type: ignore[attr-defined]
    except Exception:
        status = ""

    return {"school": school, "program": program, "intake": intake, "status": status}


def _go_requirements_tab(driver: ab.webdriver.Chrome) -> None:
    # Some pages have a "Requirements" tab; others show requirements by default.
    btn = _first_visible(
        driver,
        [
            "//*[@role='tab' and contains(., 'Requirement')]",
            "//*[@role='tab' and contains(., 'Required')]",
            "//*[self::a or self::button][contains(normalize-space(.), 'Requirements')]",
            "//*[self::a or self::button][contains(normalize-space(.), 'Required')]",
        ],
    )
    if btn is None:
        return
    try:
        ab.click_with_retry(driver, btn, timeout_s=12, retries=3)
        time.sleep(1.0)
    except Exception:
        try:
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(1.0)
        except Exception:
            return


def _scrape_requirements_table(driver: ab.webdriver.Chrome) -> List[Dict[str, str]]:
    """
    Try to scrape any visible table/grid on the requirements page.
    Output: list of dicts with generic column keys: col1, col2, ...
    """
    # Wait for something that looks like requirements content
    try:
        WebDriverWait(driver, 25).until(lambda d: "require" in ((d.find_element(By.TAG_NAME, "body").text or "").lower()))
    except Exception:
        pass

    roots: List[WebElement] = []
    for css in ["table", "[role='table']", "[role='grid']"]:
        try:
            roots.extend(driver.find_elements(By.CSS_SELECTOR, css))
        except Exception:
            continue

    def visible_rows(root: WebElement) -> List[WebElement]:
        for sel in ["tbody tr", "[role='row']"]:
            try:
                rs = root.find_elements(By.CSS_SELECTOR, sel)
                rs = [r for r in rs if _safe_text(r)]
                if rs:
                    return rs
            except Exception:
                continue
        return []

    best_root = None
    best_n = 0
    for r in roots:
        try:
            if not r.is_displayed():
                continue
        except Exception:
            pass
        rs = visible_rows(r)
        if len(rs) > best_n:
            best_n = len(rs)
            best_root = r

    if best_root is None or best_n == 0:
        # Fallback: scrape bullet lists / cards
        items: List[Dict[str, str]] = []
        try:
            for el in driver.find_elements(By.XPATH, "//*[self::li or self::p or self::div][contains(., 'Required') or contains(., 'Document') or contains(., 'Upload')]"):
                t = _safe_text(el)
                if t and len(t) <= 280:
                    items.append({"text": t})
        except Exception:
            pass
        return items[:200]

    rows = visible_rows(best_root)
    out: List[Dict[str, str]] = []
    for r in rows[:250]:
        # Skip header-ish rows
        t = _safe_text(r)
        if not t:
            continue
        if "document" in t.lower() and "status" in t.lower() and len(t) < 90:
            continue
        cells: List[str] = []
        try:
            tds = r.find_elements(By.CSS_SELECTOR, "td")
            if tds:
                cells = [(_safe_text(td) or "").replace("\n", " ").strip() for td in tds]
        except Exception:
            pass
        if not cells:
            try:
                divs = r.find_elements(By.CSS_SELECTOR, "[role='cell']")
                if divs:
                    cells = [(_safe_text(td) or "").replace("\n", " ").strip() for td in divs]
            except Exception:
                pass
        if not cells:
            cells = [t.replace("\n", " ").strip()]
        d: Dict[str, str] = {}
        for i, c in enumerate(cells):
            if c:
                d[f"col{i+1}"] = c
        if d:
            out.append(d)
    return out


def scrape_required_documents_for_student(
    driver: ab.webdriver.Chrome,
    wait: WebDriverWait,
    err_dir: Path,
    student_id: str,
) -> List[RequirementRow]:
    """
    For one student profile (assumes we are already on that student's Applications page),
    open each Paid Application 'View' page and scrape Requirements.
    """
    rows: List[RequirementRow] = []

    # Ensure we're on Applications via DIRECT URL (most stable in ApplyBoard SPA)
    base_url = None
    try:
        u = driver.current_url or ""
        m = re.search(r"(https?://[^/]+/students/\d+)", u)
        base_url = m.group(1) if m else None
    except Exception:
        base_url = None
    if base_url:
        try:
            driver.get(base_url + "/applications")
        except Exception:
            pass
    ab.wait_for_page(driver)
    time.sleep(1.0)
    try:
        ab.dismiss_applyboard_lightboxes(driver)
    except Exception:
        pass

    # Wait for paid apps signals
    try:
        WebDriverWait(driver, 90).until(
            lambda d: len(_find_paid_app_view_buttons(d)) > 0
            or "no paid applications" in ((d.find_element(By.TAG_NAME, "body").text or "").lower())
        )
    except Exception:
        pass

    view_pairs = _find_paid_app_view_buttons(driver)
    if not view_pairs:
        # Keep behavior visible in terminal via debug artifacts
        try:
            ab.screenshot_on_error(driver, err_dir, f"paid_none_{student_id}")
        except Exception:
            pass
        return rows

    # De-dup by element identity (sometimes same button appears twice in DOM)
    uniq: List[Tuple[WebElement, WebElement]] = []
    seen = set()
    for vb, row in view_pairs:
        k = id(vb)
        if k in seen:
            continue
        seen.add(k)
        uniq.append((vb, row))

    apps_url = (base_url + "/applications") if base_url else (driver.current_url or "")

    for i, (vb, row_el) in enumerate(uniq, start=1):
        meta = _extract_paid_row_metadata(row_el)
        view_url = _extract_view_url(vb)

        try:
            # Navigate to View page in SAME TAB (more stable than tab handle juggling).
            if view_url:
                driver.get(view_url)
            else:
                # Fallback: click (may navigate in same tab)
                ab.click_with_retry(driver, vb, timeout_s=12, retries=3)
            ab.wait_for_page(driver)
            time.sleep(1.2)
            _go_requirements_tab(driver)
            reqs = _scrape_requirements_table(driver)
        except Exception:
            ab.screenshot_on_error(driver, err_dir, f"req_{student_id}_{i}")
            reqs = []
        finally:
            # Always return to applications list for the next row
            try:
                if apps_url:
                    driver.get(apps_url)
                    ab.wait_for_page(driver)
                    time.sleep(1.0)
                    try:
                        ab.dismiss_applyboard_lightboxes(driver)
                    except Exception:
                        pass
            except Exception:
                pass

        rows.append(
            RequirementRow(
                student_id=student_id,
                app_index=i,
                school=meta.get("school", ""),
                program=meta.get("program", ""),
                intake=meta.get("intake", ""),
                status=meta.get("status", ""),
                requirements=reqs,
            )
        )

    return rows


def save_csv(rows: List[RequirementRow], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "student_id",
                "app_index",
                "school",
                "program",
                "intake",
                "status",
                "requirements_json",
            ],
        )
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "student_id": r.student_id,
                    "app_index": r.app_index,
                    "school": r.school,
                    "program": r.program,
                    "intake": r.intake,
                    "status": r.status,
                    "requirements_json": json.dumps(r.requirements, ensure_ascii=False),
                }
            )


def save_xlsx(rows: List[RequirementRow], out_path: Path) -> None:
    """
    Write a formatted Excel file (same styling approach as applyboard_scraper.save_results()).
    """
    import pandas as pd

    out_path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["student_id", "app_index", "school", "program", "intake", "status", "requirements_json"]
    data = []
    for r in rows:
        data.append(
            {
                "student_id": r.student_id,
                "app_index": r.app_index,
                "school": r.school,
                "program": r.program,
                "intake": r.intake,
                "status": r.status,
                "requirements_json": json.dumps(r.requirements, ensure_ascii=False),
            }
        )
    df = pd.DataFrame(data, columns=cols)

    xlsx_path = out_path
    try:
        df.to_excel(xlsx_path, index=False)
    except PermissionError:
        alt = out_path.parent / f"applyboard_required_documents_{ab.now_stamp()}.xlsx"  # type: ignore[attr-defined]
        df.to_excel(alt, index=False)
        xlsx_path = alt

    # Excel formatting (openpyxl) — keep best-effort, matching main scraper style.
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(xlsx_path)
        ws = wb.active

        header_font = Font(bold=True, size=13, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E6DEB")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_col = ws.max_column
        max_row = ws.max_row

        ws.freeze_panes = "A2"
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        wrap_headers = {"school", "program", "requirements_json"}
        wrap_col_idxs = set()
        for i, h in enumerate(cols, start=1):
            if h in wrap_headers:
                wrap_col_idxs.add(i)

        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        ws.row_dimensions[1].height = 24

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).alignment = Alignment(
                    vertical="top",
                    wrap_text=(c in wrap_col_idxs),
                )

        MIN_W = 10
        MAX_W = 60
        for c in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=c).value or "")
            if header == "requirements_json":
                ws.column_dimensions[get_column_letter(c)].width = 60
                continue
            if header in ("program", "school"):
                ws.column_dimensions[get_column_letter(c)].width = 38
                continue
            if header == "student_id":
                ws.column_dimensions[get_column_letter(c)].width = 14
                continue

            max_len = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v).splitlines()[0]
                max_len = max(max_len, len(s))
            ws.column_dimensions[get_column_letter(c)].width = min(MAX_W, max(MIN_W, max_len + 2))

        wb.save(xlsx_path)
    except Exception:
        pass


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    out_dir, err_dir = ab.ensure_dirs(base_dir)

    load_dotenv(dotenv_path=base_dir / ".env", override=False)

    p = argparse.ArgumentParser(
        description="Scrape required documents from Paid Applications (View -> Requirements tab).",
    )
    p.add_argument("--email", default="", help="ApplyBoard email (env: APPLYBOARD_EMAIL).")
    p.add_argument("--password", default="", help="ApplyBoard password (env: APPLYBOARD_PASSWORD).")
    p.add_argument("--headless", action="store_true", help="Run Chrome headless.")
    p.add_argument("--headful", action="store_true", help="Force visible Chrome (Linux).")
    p.add_argument("--max-pages", type=int, default=0, help="Max Students pages (0=all).")
    p.add_argument("--max-students", type=int, default=0, help="Max students per page (0=all).")
    p.add_argument("--only-student-id", default="", help="Scrape only this Student ID (debug).")
    p.add_argument("--login-url", default="", help="Override login entry URL (rare).")
    p.add_argument("--slow", type=float, default=0.0, help="Extra sleep seconds between major steps (debug/visibility).")
    p.add_argument("--verbose", action="store_true", help="Print each step to terminal.")
    args = p.parse_args()

    email = (args.email or os.getenv("APPLYBOARD_EMAIL") or "").strip()
    password = (args.password or os.getenv("APPLYBOARD_PASSWORD") or "").strip()

    headless = bool(args.headless)
    if sys.platform != "win32":
        headless = headless or (os.getenv("APPLYBOARD_HEADLESS", "").strip().lower() in ("1", "true", "yes", "on"))
        if sys.platform.startswith("linux") and not (os.getenv("DISPLAY") or "").strip():
            headless = True
        if os.getenv("APPLYBOARD_HEADFUL", "").strip().lower() in ("1", "true", "yes", "on"):
            headless = False
        if args.headful:
            headless = False

    if not email:
        email = input("ApplyBoard email: ").strip()
    if not password:
        password = getpass("ApplyBoard password: ")

    driver: Optional[ab.webdriver.Chrome] = None
    all_rows: List[RequirementRow] = []

    # If user didn't pass --verbose, default to verbose so progress is always visible.
    if not args.verbose:
        args.verbose = True

    try:
        driver = ab.build_driver(headless=headless)
        wait = WebDriverWait(driver, ab.DEFAULT_TIMEOUT_S)
        if args.verbose:
            print(f"Opening ApplyBoard login -> {ab.effective_login_url(args.login_url)}", flush=True)
        ab.login(driver, wait, email, password, err_dir, login_entry_url=ab.effective_login_url(args.login_url))
        if args.slow:
            time.sleep(float(args.slow))
        if args.verbose:
            print("Navigating to Students page...", flush=True)
        ab.wait_for_students_table(driver, wait, err_dir)
        if args.slow:
            time.sleep(float(args.slow))

        def wait_students_page_ready(timeout_s: int = 180) -> List[WebElement]:
            """
            Use the SAME readiness heuristics as the main scraper:
            - wait for SPA load (document.readyState)
            - wait for real student rows (numeric Student ID) OR an empty-state message.
            """
            ab.wait_for_page(driver, timeout_s=min(60, timeout_s))
            deadline = time.time() + timeout_s
            last = []
            last_beat = 0.0
            while time.time() < deadline:
                if args.verbose and (time.time() - last_beat) >= 5.0:
                    try:
                        u = driver.current_url or ""
                    except Exception:
                        u = ""
                    print(f"[WAIT] Students table loading... url={u}", flush=True)
                    last_beat = time.time()
                try:
                    ab.dismiss_applyboard_lightboxes(driver)
                except Exception:
                    pass
                try:
                    # This helper already filters to "real" rows with numeric IDs.
                    last = ab.get_student_data_rows(driver)
                    if last:
                        return last
                except Exception:
                    last = []

                try:
                    body = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
                    if "no students" in body or "no results" in body or "no records" in body:
                        return []
                except Exception:
                    pass

                time.sleep(0.6)
            return last

        page = 1
        while True:
            # Load current page student rows
            rows = wait_students_page_ready(timeout_s=200)
            if args.verbose:
                rng = None
                try:
                    rng = ab._read_pagination_range(driver)  # type: ignore[attr-defined]
                except Exception:
                    rng = None
                if rng:
                    print(f"\n=== Students page {page}: rows {rng[0]}-{rng[1]} of {rng[2]} ===", flush=True)
                else:
                    print(f"\n=== Students page {page} ===", flush=True)
                print(f"Found {len(rows)} student row(s) on this page.", flush=True)

            processed = 0
            for tr in rows:
                if args.max_students and processed >= int(args.max_students):
                    break

                # Extract student id from the row (numeric)
                sid = ""
                try:
                    txt = _safe_text(tr)
                    m = re.search(r"\b(\d{6,})\b", txt)
                    if m:
                        sid = m.group(1)
                except Exception:
                    sid = ""
                if not sid:
                    continue
                if args.only_student_id and sid != args.only_student_id.strip():
                    continue

                # Open profile in a new tab (reuse helper)
                parent_handle = driver.current_window_handle
                before = list(driver.window_handles)
                opened = False
                try:
                    if args.verbose:
                        print(f"\n[Student {sid}] Opening profile...", flush=True)
                    opened = ab._open_profile_in_new_tab(driver, tr, sid)  # type: ignore[attr-defined]
                except Exception:
                    opened = False
                if not opened:
                    if args.verbose:
                        print(f"[Student {sid}] Could not open profile tab (skipping).", flush=True)
                    continue

                new_handle = _switch_to_newest_tab(driver, before)
                if not new_handle:
                    if args.verbose:
                        print(f"[Student {sid}] New tab handle not found (skipping).", flush=True)
                    continue

                try:
                    # Navigate directly to /applications for stability
                    base_url = None
                    try:
                        u = driver.current_url or ""
                        m = re.search(r"(https?://[^/]+/students/\d+)", u)
                        base_url = m.group(1) if m else None
                    except Exception:
                        base_url = None
                    if base_url:
                        if args.verbose:
                            print(f"[Student {sid}] Opening Applications tab...", flush=True)
                        driver.get(base_url + "/applications")
                        ab.wait_for_page(driver)
                        time.sleep(1.0 + float(args.slow or 0.0))
                        if args.verbose:
                            print(f"[Student {sid}] Applications loaded. Finding paid applications...", flush=True)

                    rs = scrape_required_documents_for_student(driver, wait, err_dir, sid)
                    all_rows.extend(rs)
                    if args.verbose:
                        if not rs:
                            print(f"[Student {sid}] No paid application universities found (no View rows).", flush=True)
                        else:
                            print(f"[Student {sid}] Scraped requirements for {len(rs)} paid application(s).", flush=True)
                except Exception:
                    ab.screenshot_on_error(driver, err_dir, f"student_{sid}_req_fatal")
                    if args.verbose:
                        print(f"[Student {sid}] Error while scraping requirements (see screenshots).", flush=True)
                finally:
                    _close_tab_and_return(driver, new_handle, parent_handle)

                processed += 1

                # Save after each student (safe for long runs)
                save_csv(all_rows, out_dir / OUTPUT_CSV)
                save_xlsx(all_rows, out_dir / OUTPUT_XLSX)
                if args.verbose:
                    print(
                        f"[Saved] {len(all_rows)} application requirement row(s) -> {out_dir / OUTPUT_CSV} | {out_dir / OUTPUT_XLSX}",
                        flush=True,
                    )

            if args.only_student_id:
                break

            if args.max_pages and page >= int(args.max_pages):
                break

            # Give the SPA a moment before paginating to avoid flick/false empties.
            time.sleep(0.8 + float(args.slow or 0.0))
            if not ab.go_next_page(driver, wait, err_dir):
                if args.verbose:
                    print("No next page detected. Finished.", flush=True)
                break
            # Wait for next page to actually render rows (same logic as existing script)
            try:
                wait_students_page_ready(timeout_s=200)
            except Exception:
                pass
            page += 1

    except (TimeoutException, WebDriverException, NoSuchElementException) as exc:
        if driver is not None:
            ab.screenshot_on_error(driver, err_dir, "fatal_required_docs")
        raise exc
    finally:
        ab._safe_quit_driver(driver)  # type: ignore[attr-defined]


if __name__ == "__main__":
    main()

